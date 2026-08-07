"""Métricas de valor generado por usuario (período de suscripción actual)."""

from __future__ import annotations

import io
import logging
from dataclasses import dataclass
from datetime import date
from typing import Any, Callable

from uso_metricas import (
    METRICAS_USO,
    fila_uso_mes_desde_bucket,
    fila_uso_mes_vacia,
    mes_keys_uso,
    meta_key_a_mes_key,
    meta_keys_uso,
    metricas_dashboard_desde_uso,
)

_LOG = logging.getLogger(__name__)

_USO_KEYS = meta_keys_uso()
_USO_KEY_A_MES = meta_key_a_mes_key()
_USO_MES_CAMPOS = mes_keys_uso()

_MESES_HISTORIAL_USO = 3


def _meses_vigentes_uso(hoy: date | None = None) -> set[str]:
    """Últimos N meses calendario (incluye el mes en curso)."""
    ref = hoy or date.today()
    y, m = ref.year, ref.month
    out: set[str] = set()
    for _ in range(_MESES_HISTORIAL_USO):
        out.add(f"{y:04d}-{m:02d}")
        m -= 1
        if m <= 0:
            m = 12
            y -= 1
    return out


def _podar_uso_por_mes_meta(meta: dict[str, Any], hoy: date | None = None) -> None:
    por_mes = meta.get("uso_por_mes")
    if not isinstance(por_mes, dict) or not por_mes:
        return
    vigentes = _meses_vigentes_uso(hoy)
    meta["uso_por_mes"] = {
        mes: bucket
        for mes, bucket in por_mes.items()
        if mes in vigentes and isinstance(bucket, dict)
    }


def reset_uso_periodo_meta(meta: dict[str, Any]) -> None:
    """Reinicia contadores del período de suscripción; conserva uso mensual (3 meses)."""
    for key in _USO_KEYS:
        meta[key] = 0
    _podar_uso_por_mes_meta(meta)


def _incrementar_uso_por_mes_meta(meta: dict[str, Any], mes: str, **campos: int) -> None:
    por_mes = meta.get("uso_por_mes")
    if not isinstance(por_mes, dict):
        por_mes = {}
    bucket = por_mes.get(mes)
    if not isinstance(bucket, dict):
        bucket = {}
    for campo, val in campos.items():
        if val <= 0 or campo not in _USO_MES_CAMPOS:
            continue
        try:
            actual = int(bucket.get(campo) or 0)
        except (TypeError, ValueError):
            actual = 0
        bucket[campo] = min(actual + int(val), 1_000_000_000)
    por_mes[mes] = bucket
    meta["uso_por_mes"] = por_mes
    _podar_uso_por_mes_meta(meta)


def registrar_uso_cuit_mes_en_meta(meta: dict[str, Any], cantidad: int = 1) -> None:
    if cantidad <= 0:
        return
    mes = date.today().strftime("%Y-%m")
    _incrementar_uso_por_mes_meta(meta, mes, cuit=cantidad)


def _leer_uso_por_mes(meta: dict[str, Any]) -> list[dict[str, Any]]:
    por_mes = meta.get("uso_por_mes")
    if not isinstance(por_mes, dict):
        por_mes = {}
    filas: list[dict[str, Any]] = []
    for mes in sorted(_meses_vigentes_uso()):
        bucket = por_mes.get(mes)
        if not isinstance(bucket, dict):
            bucket = {}
        try:
            filas.append(
                fila_uso_mes_desde_bucket(mes, _formatear_mes_ym(mes), bucket)
            )
        except (TypeError, ValueError):
            filas.append(fila_uso_mes_vacia(mes, _formatear_mes_ym(mes)))
    return filas


def _formatear_mes_ym(mes: str) -> str:
    """YYYY-MM → MM/YYYY para mostrar."""
    try:
        y, m = mes.split("-", 1)
        return f"{int(m):02d}/{y}"
    except (TypeError, ValueError, IndexError):
        return mes


def _leer_uso_meta(meta: dict[str, Any]) -> dict[str, int]:
    out: dict[str, int] = {}
    for key in _USO_KEYS:
        try:
            out[key] = max(0, min(int(meta.get(key) or 0), 1_000_000_000))
        except (TypeError, ValueError):
            out[key] = 0
    return out


def _incrementar_uso(username: str, **campos: int) -> None:
    from auth import es_administrador
    from auth_registro import (
        _cargar_overlay_completo,
        _guardar_overlay_completo,
        _lock,
        meta_es_admin,
        resolver_clave_usuario_overlay,
    )

    u_raw = (username or "").strip()
    if not u_raw or es_administrador(u_raw):
        return
    u = resolver_clave_usuario_overlay(u_raw)
    if not u:
        _LOG.warning(
            "Uso no registrado: usuario %r no está en usuarios_registrados.",
            u_raw,
        )
        return
    incrementos = {k: max(0, int(v)) for k, v in campos.items() if int(v) > 0}
    if not incrementos:
        return
    with _lock:
        overlay = _cargar_overlay_completo()
        users = overlay.get("users")
        if not isinstance(users, dict) or u not in users:
            _LOG.warning(
                "Uso no registrado: clave overlay %r (desde %r) ausente en el store.",
                u,
                u_raw,
            )
            return
        meta = users[u]
        if not isinstance(meta, dict) or meta_es_admin(meta):
            return
        uso = _leer_uso_meta(meta)
        for key, val in incrementos.items():
            if key not in _USO_KEYS:
                continue
            uso[key] = min(uso[key] + val, 1_000_000_000)
        for key in _USO_KEYS:
            meta[key] = uso[key]
        mes = date.today().strftime("%Y-%m")
        mes_campos = {
            _USO_KEY_A_MES[k]: v for k, v in incrementos.items() if k in _USO_KEY_A_MES
        }
        if mes_campos:
            _incrementar_uso_por_mes_meta(meta, mes, **mes_campos)
        _guardar_overlay_completo(overlay)
    _sync_uso_remoto_si_corresponde(u, incrementos)


def _sync_uso_remoto_si_corresponde(username: str, incrementos: dict[str, int]) -> None:
    """Portable: replica métricas de uso en el servidor (Neon) para el dashboard admin."""
    if not incrementos:
        return
    try:
        from auth_registro_db import enabled

        if enabled():
            return
        from auth import _modo_remoto_activo

        if not _modo_remoto_activo():
            return
        from auth_registro import registrar_uso_remoto

        registrar_uso_remoto(username, incrementos)
    except Exception:
        _LOG.debug("Sync remoto de uso omitido", exc_info=True)


def contar_comprobantes_en_archivo(datos: bytes, nombre: str) -> int:
    try:
        from sumar_imp_total import leer_tabla

        df = leer_tabla(io.BytesIO(datos), nombre_archivo=nombre or "comprobantes.csv")
        return max(0, len(df))
    except Exception as exc:
        _LOG.debug("No se pudo contar comprobantes en %s: %s", nombre, exc)
        return 0


def contadores_mc_desde_resultado(resultado) -> tuple[int, int]:
    mce = (
        contar_comprobantes_en_archivo(resultado.emitidos[0], resultado.emitidos[1])
        if getattr(resultado, "emitidos", None)
        else 0
    )
    mcr = (
        contar_comprobantes_en_archivo(resultado.recibidos[0], resultado.recibidos[1])
        if getattr(resultado, "recibidos", None)
        else 0
    )
    return mce, mcr


def registrar_uso_mc(username: str, *, mce: int = 0, mcr: int = 0) -> None:
    _incrementar_uso(
        username,
        uso_mce_comprobantes=max(0, int(mce)),
        uso_mcr_comprobantes=max(0, int(mcr)),
    )


def registrar_uso_dfe(username: str, notificaciones: int) -> None:
    _incrementar_uso(username, uso_dfe_notificaciones=max(0, int(notificaciones)))


def registrar_uso_vl(username: str, cuits: int = 1) -> None:
    _incrementar_uso(username, uso_vl_cuits=max(0, int(cuits)))


def registrar_uso_np(username: str) -> None:
    _incrementar_uso(username, uso_np_cuits=1)


@dataclass
class RegistroValorUso:
    """Callbacks de registro de uso por servicio (extensible vía ``uso_metricas.METRICAS_USO``)."""

    usuario: str

    def mc(self, mce: int = 0, mcr: int = 0) -> None:
        registrar_uso_mc(self.usuario, mce=mce, mcr=mcr)

    def dfe(self, notificaciones: int = 0) -> None:
        registrar_uso_dfe(self.usuario, notificaciones)

    def vl(self, cuits: int = 1) -> None:
        registrar_uso_vl(self.usuario, cuits=cuits)

    def np(self) -> None:
        registrar_uso_np(self.usuario)


def fabricar_registro_valor(username: str | None) -> RegistroValorUso | None:
    u = (username or "").strip()
    if not u:
        return None
    from auth import es_administrador

    if es_administrador(u):
        return None
    return RegistroValorUso(u)


def dashboard_valor_usuario(cuit: str) -> dict[str, Any] | None:
    from auth_registro import (
        _parse_fecha_local,
        cargar_usuarios_overlay,
        formatear_cuit,
        meta_es_admin,
        normalizar_cuit,
        resolver_clave_overlay,
        _leer_cupo_meta,
    )

    u = resolver_clave_overlay(cuit) or normalizar_cuit(cuit)
    if not u:
        return None
    meta = cargar_usuarios_overlay().get(u)
    if not isinstance(meta, dict) or meta_es_admin(meta):
        return None
    if meta.get("pendiente_aprobacion"):
        return None
    from auth_registro import _telefono_desde_meta

    limite, usados = _leer_cupo_meta(meta)
    uso = _leer_uso_meta(meta)
    vd = _parse_fecha_local(meta.get("valido_desde"))
    vh = _parse_fecha_local(meta.get("valido_hasta"))
    tel = _telefono_desde_meta(meta)
    out: dict[str, Any] = {
        "cuit": u,
        "cuit_fmt": formatear_cuit(u),
        "nombre": meta.get("nombre") or "",
        "email": meta.get("email") or "",
        "telefono": tel.get("fmt") or "",
        "telefono_url": tel.get("url") or "",
        "valido_desde_fmt": vd.strftime("%d/%m/%Y") if vd else "—",
        "valido_hasta_fmt": vh.strftime("%d/%m/%Y") if vh else "—",
        "cuit_usados": usados,
        "cuit_limite": limite,
        "cuit_disponibles": max(0, limite - usados),
        "uso_por_mes": _leer_uso_por_mes(meta),
        "metricas_uso": METRICAS_USO,
    }
    out.update(metricas_dashboard_desde_uso(uso))
    return out


def listar_dashboards_valor() -> list[dict[str, Any]]:
    from auth_registro import listar_usuarios_suscripcion

    out: list[dict[str, Any]] = []
    for sub in listar_usuarios_suscripcion():
        cuit = sub.get("cuit") or ""
        try:
            dash = dashboard_valor_usuario(cuit)
        except Exception:
            _LOG.exception("No se pudo armar dashboard de valor para %r", cuit)
            continue
        if dash:
            out.append(dash)
    out.sort(key=lambda d: (d.get("cuit_fmt") or d.get("cuit") or "").lower())
    return out


def _nombre_hoja_excel(d: dict[str, Any], usados: set[str]) -> str:
    import re

    base = re.sub(r"[\\/*?:\[\]]", "", (d.get("cuit_fmt") or d.get("cuit") or "Usuario")).strip()
    base = base.replace("—", "-")[:28] or "Usuario"
    nombre = base
    n = 1
    while nombre in usados:
        suf = f"_{n}"
        nombre = f"{base[: max(1, 31 - len(suf))]}{suf}"
        n += 1
    usados.add(nombre)
    return nombre


def _valor_celda_excel(val: Any) -> Any:
    """Quita caracteres de control que openpyxl/Excel no admiten."""
    if val is None or isinstance(val, (int, float, bool)):
        return val
    if not isinstance(val, str):
        val = str(val)
    import re

    return re.sub(r"[\000-\010\013\014\016-\037]", "", val)


def _hipervinculo_hoja(hoja: str) -> str:
    """Referencia interna Excel; escapa comillas simples del nombre de hoja."""
    safe = (hoja or "Usuario").replace("'", "''")
    return f"#'{safe}'!A1"


def generar_excel_dashboard_valor(
    dashboards: list[dict[str, Any]] | None = None,
) -> bytes:
    """Excel: hoja Resumen (filtros + enlaces), Uso por mes y una hoja de detalle por usuario."""
    from datetime import date

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    if dashboards is None:
        dashboards = listar_dashboards_valor()
    wb = Workbook()
    ws = wb.active
    ws.title = "Resumen"

    encabezados = [
        "CUIT / Usuario",
        "Nombre",
        "Email",
        "Telefono",
        "Valido desde",
        "Valido hasta",
        "CUIT procesados",
        "Cupo limite",
        "CUIT disponibles",
    ] + [m.excel_label for m in METRICAS_USO]
    ws.append(encabezados)
    for col in range(1, len(encabezados) + 1):
        c = ws.cell(row=1, column=col)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="1F4E79")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    hojas_usuario: list[tuple[str, dict[str, Any]]] = []
    usados: set[str] = {"Resumen", "Uso por mes"}
    for d in dashboards:
        hojas_usuario.append((_nombre_hoja_excel(d, usados), d))

    fila = 2
    for hoja, d in hojas_usuario:
        fila_datos = [
            _valor_celda_excel(d.get("cuit_fmt") or d.get("cuit")),
            _valor_celda_excel(d.get("nombre") or ""),
            _valor_celda_excel(d.get("email") or ""),
            _valor_celda_excel(d.get("telefono") or ""),
            _valor_celda_excel(d.get("valido_desde_fmt") or ""),
            _valor_celda_excel(d.get("valido_hasta_fmt") or ""),
            int(d.get("cuit_usados", 0) or 0),
            int(d.get("cuit_limite", 0) or 0),
            int(d.get("cuit_disponibles", 0) or 0),
        ] + [int(d.get(m.dash_key, 0) or 0) for m in METRICAS_USO]
        ws.append(fila_datos)
        celda = ws.cell(row=fila, column=1)
        try:
            celda.hyperlink = _hipervinculo_hoja(hoja)
            celda.font = Font(color="0563C1", underline="single")
        except Exception:
            pass
        fila += 1

    ultima_fila = max(1, len(dashboards) + 1)
    ultima_col = get_column_letter(len(encabezados))
    if dashboards:
        ws.auto_filter.ref = f"A1:{ultima_col}{ultima_fila}"

    ws.freeze_panes = "A2"
    for col in range(1, len(encabezados) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 16
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 26
    ws.column_dimensions["D"].width = 16

    col_ayuda = get_column_letter(len(encabezados) + 2)
    ws[f"{col_ayuda}1"] = "Instrucciones"
    ws[f"{col_ayuda}2"] = (
        "Hace clic en el CUIT de la tabla para abrir el detalle del usuario en otra hoja. "
        "La hoja Uso por mes concentra el desglose mensual por sistema. "
        "Podes filtrar y ordenar con los controles de la tabla."
    )
    ws[f"{col_ayuda}2"].alignment = Alignment(wrap_text=True)
    ws.column_dimensions[col_ayuda].width = 36

    titulo_fill = PatternFill("solid", fgColor="E8F0FE")
    lbl_font = Font(bold=True)
    enc_mes = ["Mes", "CUIT procesados"] + [m.excel_label for m in METRICAS_USO]

    ws_mes = wb.create_sheet(title="Uso por mes")
    enc_mes_resumen = ["CUIT / Usuario", "Nombre"] + enc_mes
    ws_mes.append(enc_mes_resumen)
    for col in range(1, len(enc_mes_resumen) + 1):
        c = ws_mes.cell(row=1, column=col)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="1F4E79")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    fila_mes_global = 2
    n_cols_mes = len(enc_mes_resumen)
    for _hoja, d in hojas_usuario:
        uso_mes = d.get("uso_por_mes") or []
        if not isinstance(uso_mes, list):
            continue
        for fila_mes in uso_mes:
            if not isinstance(fila_mes, dict):
                continue
            ws_mes.append(
                [
                    _valor_celda_excel(d.get("cuit_fmt") or d.get("cuit")),
                    _valor_celda_excel(d.get("nombre") or ""),
                    _valor_celda_excel(fila_mes.get("mes_fmt") or fila_mes.get("mes")),
                    int(fila_mes.get("cuit", 0) or 0),
                ]
                + [int(fila_mes.get(m.mes_key, 0) or 0) for m in METRICAS_USO]
            )
            fila_mes_global += 1
    if fila_mes_global == 2:
        ws_mes.append(
            ["-", "-", "Sin datos mensuales en los ultimos 3 meses."]
            + [""] * (n_cols_mes - 3)
        )
    ultima_fila_mes = max(1, ws_mes.max_row)
    ultima_col_mes = get_column_letter(n_cols_mes)
    if fila_mes_global > 2:
        ws_mes.auto_filter.ref = f"A1:{ultima_col_mes}{ultima_fila_mes}"
    ws_mes.freeze_panes = "A2"
    for col in range(1, n_cols_mes + 1):
        ws_mes.column_dimensions[get_column_letter(col)].width = 18 if col > 2 else 20

    filas_detalle = [
        ("CUIT / Usuario", lambda d: d.get("cuit_fmt") or d.get("cuit")),
        ("Nombre", lambda d: d.get("nombre") or "-"),
        ("Email", lambda d: d.get("email") or "-"),
        ("Telefono", lambda d: d.get("telefono") or "-"),
        ("Periodo desde", lambda d: d.get("valido_desde_fmt") or "-"),
        ("Periodo hasta", lambda d: d.get("valido_hasta_fmt") or "-"),
        ("CUIT procesados (cupo)", lambda d: d.get("cuit_usados", 0)),
        ("Cupo limite", lambda d: d.get("cuit_limite", 0)),
        ("CUIT disponibles", lambda d: d.get("cuit_disponibles", 0)),
    ] + [
        (m.excel_label, lambda d, _m=m: d.get(_m.dash_key, 0))
        for m in METRICAS_USO
    ]

    def _escribir_tabla_uso_mes(hoja_ws, fila_inicio: int, uso_mes: list[dict[str, Any]]) -> int:
        r = fila_inicio
        for col, tit in enumerate(enc_mes, start=1):
            c = hoja_ws.cell(row=r, column=col, value=tit)
            c.font = lbl_font
            c.fill = titulo_fill
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        r += 1
        n_cols = len(enc_mes)
        if uso_mes:
            for fila_mes in uso_mes:
                if not isinstance(fila_mes, dict):
                    continue
                hoja_ws.cell(
                    row=r,
                    column=1,
                    value=_valor_celda_excel(fila_mes.get("mes_fmt") or fila_mes.get("mes")),
                )
                hoja_ws.cell(row=r, column=2, value=int(fila_mes.get("cuit", 0) or 0))
                for idx, m in enumerate(METRICAS_USO, start=3):
                    hoja_ws.cell(row=r, column=idx, value=int(fila_mes.get(m.mes_key, 0) or 0))
                r += 1
        else:
            hoja_ws.cell(
                row=r,
                column=1,
                value="Sin datos mensuales en los ultimos 3 meses.",
            )
            hoja_ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=n_cols)
            r += 1
        return r

    for hoja, d in hojas_usuario:
        try:
            det = wb.create_sheet(title=hoja)
            titulo = (
                "Dashboard de valor - "
                + str(_valor_celda_excel(d.get("cuit_fmt") or d.get("cuit") or ""))
            )
            det["A1"] = titulo
            det["A1"].font = Font(bold=True, size=14)
            det["A2"] = "Volver al resumen"
            try:
                det["A2"].hyperlink = "#Resumen!A1"
                det["A2"].font = Font(color="0563C1", underline="single")
            except Exception:
                pass
            det["A4"] = "Metrica"
            det["B4"] = "Valor"
            det["A4"].font = lbl_font
            det["B4"].font = lbl_font
            det["A4"].fill = titulo_fill
            det["B4"].fill = titulo_fill
            r = 5
            for etiqueta, fn in filas_detalle:
                try:
                    valor = fn(d)
                except Exception:
                    valor = ""
                det.cell(row=r, column=1, value=_valor_celda_excel(etiqueta))
                det.cell(row=r, column=2, value=_valor_celda_excel(valor))
                r += 1

            r += 2
            titulo_mes = r
            det.cell(row=titulo_mes, column=1, value="Uso por mes")
            det.cell(row=titulo_mes, column=1).font = Font(bold=True, size=12)
            uso_mes = d.get("uso_por_mes") or []
            if not isinstance(uso_mes, list):
                uso_mes = []
            r = _escribir_tabla_uso_mes(det, titulo_mes + 1, uso_mes)

            det.column_dimensions["A"].width = 32
            det.column_dimensions["B"].width = 22
            for col in range(3, len(enc_mes) + 1):
                det.column_dimensions[get_column_letter(col)].width = 18
            det["D4"] = f"Exportado: {date.today().strftime('%d/%m/%Y')}"
        except Exception:
            _LOG.exception("No se pudo armar hoja Excel para %r", hoja)
            continue

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
