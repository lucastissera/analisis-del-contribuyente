#!/usr/bin/env python3
"""
Ingeniería inversa de PDF de liquidaciones (certificado / LPG / hacienda).
Genera JSON estructurado + Excel comparativo por archivo analizado.
"""

from __future__ import annotations

import json
import re
import sys
import time
from pathlib import Path
from typing import Any

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Font

from liquidaciones_pdf import (
    MSG_AJUSTE_UNIFICADO,
    _COLS_CERTIFICADOS,
    _COLS_HACIENDA,
    _COLS_LPG,
    _clasificar_pdf,
    _leer_texto_pdf,
    _paginas_hacienda,
    _parse_numero,
    escribir_excel_resumen,
    extraer_certificado_deposito,
    extraer_hacienda_pagina,
    extraer_lpg,
    procesar_pdf,
)

PDFS_DEFAULT = [
    Path(
        r"d:\Lucas\Cursor Programacion\zAnalisis Integral del Contribuyente"
        r"\Proceso liquidaciones\Para analizar\BOCA DEL RIO S.A"
        r"\Hacienda\Receptor\00002 - 00000058.pdf"
    ),
    Path(
        r"d:\Lucas\Cursor Programacion\zAnalisis Integral del Contribuyente"
        r"\Proceso liquidaciones\Para analizar\BOCA DEL RIO S.A"
        r"\Primarias\Recibidas\330129928007.pdf"
    ),
    Path(
        r"d:\Lucas\Cursor Programacion\zAnalisis Integral del Contribuyente"
        r"\Proceso liquidaciones\Para analizar\BOCA DEL RIO S.A"
        r"\Primarias\Recibidas\330229379940.pdf"
    ),
    Path(
        r"d:\Lucas\Cursor Programacion\zAnalisis Integral del Contribuyente"
        r"\Proceso liquidaciones\Para analizar\BOCA DEL RIO S.A"
        r"\Primarias\Recibidas\330229430049.pdf"
    ),
]


def _buscar(texto: str, patron: str, *, flags: int = re.I) -> str | None:
    m = re.search(patron, texto, flags)
    return m.group(1).strip() if m else None


def _num(val: str | None) -> float | None:
    return _parse_numero(val)


def _aplanar(d: Any, prefijo: str = "") -> dict[str, Any]:
    out: dict[str, Any] = {}
    if isinstance(d, dict):
        for k, v in d.items():
            clave = f"{prefijo}.{k}" if prefijo else str(k)
            out.update(_aplanar(v, clave))
    elif isinstance(d, list):
        out[prefijo or "lista"] = json.dumps(d, ensure_ascii=False, default=str)
    else:
        out[prefijo or "valor"] = d
    return out


def _secciones_lpg(texto: str, paginas: list[str]) -> list[dict[str, Any]]:
    secciones: list[dict[str, Any]] = []
    patrones = [
        ("liquidacion_original", r"LIQUIDACI[^\n]*ORIGINAL"),
        ("nota_debito", r"NOTA\s+DE\s+D[EÉ]BITO"),
        ("nota_credito", r"NOTA\s+DE\s+CR[EÉ]DITO"),
        ("ajuste_unificado", r"AJUSTE\s+UNIFICADO"),
    ]
    for i, pag in enumerate(paginas):
        t = pag.upper()
        for clave, pat in patrones:
            if re.search(pat, t, re.I):
                secciones.append({"pagina": i + 1, "tipo": clave, "fragmento": pag[:400]})
    if not secciones:
        tipo_doc = _buscar(texto, r"(Liquidaci[oó]n primaria de granos[^\n]*)")
        secciones.append({"pagina": 1, "tipo": "documento_unico", "titulo": tipo_doc})
    return secciones


def _parsear_paginas_lpg_detalle(paginas: list[str]) -> list[dict[str, Any]]:
    out: list[dict[str, Any]] = []
    for i, pag in enumerate(paginas):
        t = pag.upper()
        tipo_pag = "desconocido"
        if "MERCADERIA ENTREGADA" in t and "LIQUIDACION ORIGINAL" in t.replace("Ó", "O"):
            tipo_pag = "liquidacion_original"
        elif "AJUSTE D" in t and "BITO" in t:
            tipo_pag = "ajuste_debito"
        elif "AJUSTE CR" in t and "DITO" in t:
            tipo_pag = "ajuste_credito"
        elif "AJUSTE UNIFICADO" in t:
            tipo_pag = "ajuste_unificado_resumen"
        elif "AJUSTE UNIFICADO" in t or "AJUSTE CR" in t:
            tipo_pag = "ajuste"

        item: dict[str, Any] = {"pagina": i + 1, "tipo": tipo_pag}
        item["coe_original"] = _buscar(pag, r"COE ORIGINAL:\s*(\d+)")
        item["tipo_operacion"] = _buscar(pag, r"Tipo de operaci[oó]n:\s*([^\n]+)")

        m_merc = re.search(
            r"N[ºo] de Comprobante Grado Factor Contenido Proteico Peso Kg.*?\n"
            r"(\d+)\s+(G\d+)\s+(\d+)\s+([\d.,]+)\s+(\d+)",
            pag,
            re.I | re.S,
        )
        if m_merc:
            item["mercaderia_original"] = {
                "comprobante": m_merc.group(1),
                "grado": m_merc.group(2),
                "factor": m_merc.group(3),
                "cont_proteico": _num(m_merc.group(4)),
                "peso_kg": _num(m_merc.group(5)),
            }

        m_op = re.search(
            r"Cantidad Precio/Kg Subtotal.*?Operaci[^\n]*\n"
            r"([\d.,]+)\s+Kg\s+\$?([\d.,]+)\s+\$?([\d.,]+)\s+([\d.,]+)\s+\$?([\d.,]+)\s+\$?([\d.,]+)",
            pag,
            re.I | re.S,
        )
        if m_op:
            item["operacion"] = {
                "cantidad_kg": _num(m_op.group(1)),
                "precio_kg": _num(m_op.group(2)),
                "subtotal": _num(m_op.group(3)),
                "alicuota_iva_pct": _num(m_op.group(4)),
                "importe_iva": _num(m_op.group(5)),
                "operacion_con_iva": _num(m_op.group(6)),
            }

        item["importe_neto_pagar"] = _num(
            _buscar(pag, r"Importe Neto a Pagar\s*\$\s*([\d.,]+)")
        )
        if tipo_pag == "ajuste_unificado_resumen":
            item["subtotal_general"] = _num(_buscar(pag, r"Subtotal General\s*\$\s*([\d.,]+)"))
            item["importe_neto"] = _num(_buscar(pag, r"Importe Neto\s*\$\s*([\d.,]+)"))
            item["pago_segun_condiciones"] = _num(
                _buscar(pag, r"Pago seg[uú]n condiciones\s*\$\s*([\d.,]+)")
            )
        if item["tipo"] != "desconocido" or any(
            k in item for k in ("operacion", "mercaderia_original", "importe_neto")
        ):
            out.append(item)
    return out


def extraer_json_lpg(texto: str, paginas: list[str]) -> dict[str, Any]:
    secciones = _secciones_lpg(texto, paginas)
    tipos = {s["tipo"] for s in secciones}
    es_ajuste = "ajuste_unificado" in tipos or "AJUSTE UNIFICADO" in texto.upper()

    pagina0 = paginas[0] if paginas else texto
    comprador = _buscar(
        pagina0,
        r"COMPRADOR\s+VENDEDOR\s+Raz[oó]n Social:\s*([^\n]+?)\s+Raz[oó]n Social:",
    )
    vendedor = _buscar(
        pagina0,
        r"COMPRADOR\s+VENDEDOR\s+Raz[oó]n Social:[^\n]+\s+Raz[oó]n Social:\s*([^\n]+)",
    )
    m_cuit = re.search(
        r"COMPRADOR\s+VENDEDOR.*?C\.U\.I\.T\.:\s*(\d+).*?C\.U\.I\.T\.:\s*(\d+)",
        pagina0,
        re.I | re.S,
    )
    coes_rel: list[str] = []
    m_coes = re.search(r"COES RELACIONADOS\s*(.*?)\s*MERCADERIA", pagina0, re.S | re.I)
    if m_coes:
        coes_rel = re.findall(r"\b(\d{10,15})\b", m_coes.group(1))

    operaciones: list[dict[str, Any]] = []
    for i, pag in enumerate(paginas):
        if not re.search(r"Cantidad Precio/Kg Subtotal", pag, re.I):
            continue
        m_op = re.search(
            r"Cantidad Precio/Kg Subtotal.*?Operaci[^\n]*\n"
            r"([\d.,]+)\s+Kg\s+\$?([\d.,]+)\s+\$?([\d.,]+)\s+([\d.,]+)\s+\$?([\d.,]+)\s+\$?([\d.,]+)",
            pag,
            re.I | re.S,
        )
        if m_op:
            operaciones.append(
                {
                    "pagina": i + 1,
                    "cantidad_kg": _num(m_op.group(1)),
                    "precio_kg": _num(m_op.group(2)),
                    "subtotal": _num(m_op.group(3)),
                    "alicuota_iva_pct": _num(m_op.group(4)),
                    "importe_iva": _num(m_op.group(5)),
                    "operacion_con_iva": _num(m_op.group(6)),
                }
            )

    return {
        "coe": _buscar(texto, r"C\.O\.E\.:\s*(\d+)"),
        "fecha_cabecera": _buscar(pagina0, r"^(\d{2}/\d{2}/\d{4})", flags=re.M),
        "comprador": {"razon_social": comprador, "cuit": m_cuit.group(1) if m_cuit else None},
        "vendedor": {"razon_social": vendedor, "cuit": m_cuit.group(2) if m_cuit else None},
        "coes_relacionados": coes_rel,
        "secciones_detectadas": secciones,
        "paginas_detalle": _parsear_paginas_lpg_detalle(paginas),
        "es_ajuste_unificado": es_ajuste,
        "operaciones_por_pagina": operaciones,
        "extraccion_excel_actual": extraer_lpg(texto, paginas),
    }


def extraer_json_hacienda(texto: str, paginas: list[str]) -> dict[str, Any]:
    paginas_usadas = _paginas_hacienda(paginas)
    bloques: list[dict[str, Any]] = []
    filas_excel: list[dict[str, Any]] = []

    pag0 = paginas[paginas_usadas[0]] if paginas_usadas else texto
    emisor = {}
    m_em = re.search(
        r"^([^\n]+)\nFecha\s+(\d{2}/\d{2}/\d{4})\nCUIT:\s*(\d+)",
        pag0,
        re.M,
    )
    if m_em:
        emisor = {
            "razon_social": m_em.group(1).strip(),
            "fecha": m_em.group(2),
            "cuit": m_em.group(3),
        }

    for idx in paginas_usadas:
        cab, animales = extraer_hacienda_pagina(paginas[idx])
        bloques.append(
            {
                "pagina_pdf": idx + 1,
                "cabecera": cab,
                "animales": animales,
                "cantidad_animales": len(animales),
            }
        )
        from liquidaciones_pdf import _filas_hacienda_excel

        filas_excel.extend(_filas_hacienda_excel(cab, animales))

    detalle_lineas: list[dict[str, Any]] = []
    m_sec = re.search(
        r"Categor[ií]a / Raza.*?Importe Bruto:",
        pag0,
        re.I | re.S,
    )
    if m_sec:
        bloque = pag0[m_sec.start() : m_sec.end()]
        for ln in bloque.splitlines():
            ln = ln.strip()
            if re.match(r"^Bovino", ln, re.I):
                detalle_lineas.append({"tipo": "categoria_inicio", "texto": ln})
            elif re.match(r"^\d+\s+Kg", ln, re.I) or re.match(r"^/", ln):
                detalle_lineas.append({"tipo": "detalle_fila", "texto": ln})

    return {
        "paginas_total": len(paginas),
        "paginas_procesadas": [i + 1 for i in paginas_usadas],
        "regla_paginacion": "3 pág→solo 1; 6 pág→1 y 2; múltiplos de 3→cada bloque de 3",
        "emisor": emisor,
        "liquidaciones": bloques,
        "detalle_animales_texto": detalle_lineas,
        "parser_animales_ok": any(b.get("cantidad_animales", 0) > 0 for b in bloques),
        "extraccion_excel_actual": filas_excel,
    }


def extraer_json_certificado(texto: str, paginas: list[str]) -> dict[str, Any]:
    import importlib.util

    mod_path = ROOT / "tools" / "_analizar_cert_pdf.py"
    spec = importlib.util.spec_from_file_location("analizar_cert_pdf", mod_path)
    if spec is None or spec.loader is None:
        return {"extraccion_excel_actual": extraer_certificado_deposito(texto)}
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    return mod.extraer_json_completo(texto, paginas)


def _timing(fn, *args, **kwargs) -> tuple[Any, float]:
    t0 = time.perf_counter()
    out = fn(*args, **kwargs)
    return out, round((time.perf_counter() - t0) * 1000, 1)


def analizar_pdf(pdf: Path) -> dict[str, Any]:
    if not pdf.is_file():
        return {"archivo": pdf.name, "error": "no_encontrado", "ruta": str(pdf)}

    salida = pdf.parent / f"{pdf.stem}_analisis"
    salida.mkdir(parents=True, exist_ok=True)

    texto, paginas = _leer_texto_pdf(pdf)
    (salida / "texto_completo.txt").write_text(texto, encoding="utf-8")

    tablas_meta: list[dict] = []
    with pdfplumber.open(pdf) as doc:
        for i, page in enumerate(doc.pages):
            tables = page.extract_tables() or []
            tablas_meta.append({"pagina": i + 1, "tablas": len(tables)})

    clasif = _clasificar_pdf(texto, paginas)

    _, ms_procesar = _timing(procesar_pdf, pdf)
    try:
        tipo, datos_actual = procesar_pdf(pdf)
        error_actual = None
    except Exception as exc:
        tipo, datos_actual = clasif, None
        error_actual = str(exc)

    if clasif == "certificado":
        json_full, ms_json = _timing(extraer_json_certificado, texto, paginas)
    elif clasif == "lpg":
        json_full, ms_json = _timing(extraer_json_lpg, texto, paginas)
    elif clasif == "hacienda":
        json_full, ms_json = _timing(extraer_json_hacienda, texto, paginas)
    else:
        json_full, ms_json = {"texto_preview": texto[:1500]}, 0.0

    payload = {
        "archivo": pdf.name,
        "ruta": str(pdf),
        "tipo_detectado": clasif,
        "paginas": len(paginas),
        "tablas_pdfplumber": tablas_meta,
        "tiempos_ms": {
            "procesar_pdf_actual": ms_procesar,
            "json_ingenieria_inversa": ms_json,
        },
        "procesar_pdf_actual": {
            "ok": error_actual is None,
            "error": error_actual,
            "tipo": tipo,
            "datos": datos_actual,
        },
        "ingenieria_inversa": json_full,
    }

    stem = pdf.stem.replace(" ", "_")
    json_path = salida / f"{stem}.json"
    json_path.write_text(
        json.dumps(payload, indent=2, ensure_ascii=False, default=str),
        encoding="utf-8",
    )

    certificados, lpg, hacienda = [], [], []
    datos_excel = json_full.get("extraccion_excel_actual")
    if clasif == "certificado" and datos_excel:
        certificados = [datos_excel]
    elif clasif == "lpg" and datos_excel:
        lpg = [datos_excel] if isinstance(datos_excel, dict) else []
    elif clasif == "hacienda" and datos_excel:
        hacienda = datos_excel if isinstance(datos_excel, list) else []

    excel_vl = salida / f"{stem}_resumen_VL.xlsx"
    escribir_excel_resumen(excel_vl, certificados=certificados, lpg=lpg, hacienda=hacienda)

    excel_cmp = salida / f"{stem}_analisis.xlsx"
    _escribir_excel_comparativo(excel_cmp, payload, clasif)

    payload["salida"] = {
        "carpeta": str(salida),
        "json": str(json_path),
        "excel_vl": str(excel_vl),
        "excel_analisis": str(excel_cmp),
    }
    return payload


def _escribir_excel_comparativo(ruta: Path, payload: dict[str, Any], clasif: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Resumen analisis"
    ws.append(["Campo", "Valor"])
    for c in ws[1]:
        c.font = Font(bold=True)
    resumen = {
        "archivo": payload.get("archivo"),
        "tipo": clasif,
        "paginas": payload.get("paginas"),
        "tiempo_procesar_ms": payload["tiempos_ms"]["procesar_pdf_actual"],
        "tiempo_json_ms": payload["tiempos_ms"]["json_ingenieria_inversa"],
        "procesar_ok": payload["procesar_pdf_actual"]["ok"],
    }
    for k, v in resumen.items():
        ws.append([k, v])

    ws2 = wb.create_sheet("JSON aplanado")
    ws2.append(["Campo", "Valor"])
    for c in ws2[1]:
        c.font = Font(bold=True)
    inv = payload.get("ingenieria_inversa") or {}
    plano = _aplanar({k: v for k, v in inv.items() if k != "extraccion_excel_actual"})
    for k in sorted(plano.keys()):
        ws2.append([k, plano[k]])

    ws3 = wb.create_sheet("Excel actual")
    datos = inv.get("extraccion_excel_actual")
    if clasif == "certificado" and datos:
        ws3.append(list(_COLS_CERTIFICADOS))
        from liquidaciones_pdf import _fila_certificado_a_lista

        ws3.append(_fila_certificado_a_lista(datos))
    elif clasif == "lpg" and datos:
        ws3.append(list(_COLS_LPG))
        ws3.append([datos.get(c) for c in _COLS_LPG])
        if datos.get("Cantidad") == MSG_AJUSTE_UNIFICADO:
            ws3.append(["NOTA", MSG_AJUSTE_UNIFICADO])
    elif clasif == "hacienda" and datos:
        ws3.append(list(_COLS_HACIENDA))
        for fila in datos:
            ws3.append([fila.get(c) for c in _COLS_HACIENDA])

    for row in ws3[1]:
        row.font = Font(bold=True)

    ws4 = wb.create_sheet("Conclusion")
    ws4.append(["Aspecto", "Detalle"])
    for c in ws4[1]:
        c.font = Font(bold=True)
    t_proc = payload["tiempos_ms"]["procesar_pdf_actual"]
    t_json = payload["tiempos_ms"]["json_ingenieria_inversa"]
    mas_rapido = "procesar_pdf (actual)" if t_proc <= t_json else "JSON inverso"
    ws4.append(["Mas rapido en este PDF", mas_rapido])
    ws4.append(
        [
            "Recomendacion",
            "Mantener pdfplumber+regex integrado; JSON inverso sirve para validar/ampliar campos, no reemplaza el motor.",
        ]
    )
    ws4.append(
        [
            "Ventaja JSON",
            "Documenta secciones (ajuste unificado, animales, COES) y facilita debug; no acelera el batch.",
        ]
    )
    wb.save(ruta)


def main(argv: list[str] | None = None) -> int:
    pdfs = [Path(a) for a in (argv or [])] if argv else PDFS_DEFAULT
    resultados: list[dict[str, Any]] = []
    for pdf in pdfs:
        print(f"\n=== {pdf.name} ===")
        res = analizar_pdf(pdf)
        if res.get("error"):
            print("ERROR:", res["error"])
            resultados.append(res)
            continue
        print("Tipo:", res["tipo_detectado"], "| Páginas:", res["paginas"])
        print("Tiempos ms:", res["tiempos_ms"])
        print("Salida:", res["salida"]["carpeta"])
        resultados.append(
            {
                "archivo": res["archivo"],
                "tipo": res["tipo_detectado"],
                "paginas": res["paginas"],
                "tiempos_ms": res["tiempos_ms"],
                "procesar_ok": res["procesar_pdf_actual"]["ok"],
                "salida": res["salida"],
            }
        )

    resumen_path = pdfs[0].parent.parent.parent / "resumen_analisis_pdfs.json"
    try:
        resumen_path.write_text(
            json.dumps(resultados, indent=2, ensure_ascii=False, default=str),
            encoding="utf-8",
        )
        print("\nResumen batch:", resumen_path)
    except Exception:
        pass
    return 0


if __name__ == "__main__":
    args = sys.argv[1:]
    raise SystemExit(main(args or None))
