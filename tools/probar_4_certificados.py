"""Prueba local de extracción de certificados (4 PDF de muestra)."""

from __future__ import annotations

import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from openpyxl import load_workbook

from liquidaciones_pdf import _COLS_CERTIFICADOS, escribir_excel_resumen, procesar_pdf

CARPETA = Path(
    r"D:\Lucas\Cursor Programacion\zAnalisis Integral del Contribuyente\Proceso liquidaciones"
)
PDFS = [
    ("TRANSFERENCIA", CARPETA / "332020806883 transferencia.pdf"),
    ("ANULADO", CARPETA / "332021206503 anulado.pdf"),
    ("RETIRO", CARPETA / "332021223502 retiro.pdf"),
    ("PLANTA", CARPETA / "Certificado deposito 332022988431.pdf"),
]
SALIDA = CARPETA / "_prueba_4_certificados.xlsx"


def main() -> None:
    filas = []
    checks: list[tuple[str, str, bool, object]] = []

    for etiqueta, pdf in PDFS:
        if not pdf.is_file():
            print(f"FALTA: {pdf.name}")
            continue
        tipo, datos = procesar_pdf(pdf)
        filas.append(datos)
        checks.append((etiqueta, "tipo_pdf", tipo == "certificado", tipo))

        if etiqueta == "TRANSFERENCIA":
            checks.append(
                (
                    etiqueta,
                    "receptor_cuit",
                    datos.get("CUIT Receptor") == 30503508725,
                    datos.get("CUIT Receptor"),
                )
            )
            checks.append(
                (
                    etiqueta,
                    "receptor_rs",
                    "MONSANTO" in (datos.get("Razon Social Receptor") or ""),
                    datos.get("Razon Social Receptor"),
                )
            )
        elif etiqueta == "ANULADO":
            checks.append(
                (
                    etiqueta,
                    "anulado_peso_vacio",
                    datos.get("Peso Bruto") is None,
                    datos.get("Peso Bruto"),
                )
            )
            checks.append(
                (
                    etiqueta,
                    "anulado_total_vacio",
                    datos.get("Total") is None,
                    datos.get("Total"),
                )
            )
            checks.append(
                (
                    etiqueta,
                    "dep_ok",
                    datos.get("Depositario CUIT") == 30502874353,
                    datos.get("Depositario CUIT"),
                )
            )
        elif etiqueta == "RETIRO":
            pb = datos.get("Peso Bruto")
            checks.append(
                (etiqueta, "peso_negativo", pb is not None and pb < 0, pb)
            )
            checks.append(
                (
                    etiqueta,
                    "peso_neto_vacio",
                    datos.get("Peso Neto") is None,
                    datos.get("Peso Neto"),
                )
            )
            checks.append(
                (
                    etiqueta,
                    "total_vacio",
                    datos.get("Total") is None,
                    datos.get("Total"),
                )
            )
        elif etiqueta == "PLANTA":
            checks.append(
                (
                    etiqueta,
                    "peso_positivo",
                    (datos.get("Peso Bruto") or 0) > 0,
                    datos.get("Peso Bruto"),
                )
            )
            checks.append(
                (
                    etiqueta,
                    "peso_neto_ok",
                    (datos.get("Peso Neto") or 0) > 0,
                    datos.get("Peso Neto"),
                )
            )

    escribir_excel_resumen(SALIDA, certificados=filas)

    print("RESULTADOS POR PDF")
    print("=" * 70)
    for (etiqueta, pdf), datos in zip(PDFS, filas):
        print(f"\n[{etiqueta}] {pdf.name}")
        print(f"  COE: {datos.get('COE')} | Tipo: {datos.get('Tipo certificado')}")
        print(
            f"  Depositario: {datos.get('Depositario Razon Social')} "
            f"({datos.get('Depositario CUIT')})"
        )
        print(
            f"  Depositante: {datos.get('Depositante Razon social')} "
            f"({datos.get('Depositante CUIT')})"
        )
        print(
            f"  Receptor: {datos.get('Razon Social Receptor')} "
            f"({datos.get('CUIT Receptor')})"
        )
        print(
            f"  Peso Bruto: {datos.get('Peso Bruto')} | "
            f"Peso Neto: {datos.get('Peso Neto')} | Total: {datos.get('Total')}"
        )

    print("\n\nVALIDACIONES")
    print("=" * 70)
    all_ok = True
    for etiqueta, test, ok, val in checks:
        mark = "OK" if ok else "FALLO"
        if not ok:
            all_ok = False
        print(f"[{mark}] {etiqueta} / {test}: {val!r}")

    wb = load_workbook(SALIDA, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    headers = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
    print(f"\nExcel: {SALIDA.name} | hoja: {ws.title} | filas: {ws.max_row - 1}")
    print(f"Columnas coinciden con plantilla: {headers == _COLS_CERTIFICADOS}")
    print(f"\nRESULTADO GLOBAL: {'OK' if all_ok else 'CON FALLOS'}")


if __name__ == "__main__":
    main()
