#!/usr/bin/env python3
"""Ingeniería inversa: certificación electrónica de granos → JSON + Excel."""

from __future__ import annotations

import json
import re
import sys
from pathlib import Path
from typing import Any

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from openpyxl import Workbook
from openpyxl.styles import Font

from liquidaciones_pdf import (
    _COLS_CERTIFICADOS,
    _leer_texto_pdf,
    _parse_numero,
    escribir_excel_resumen,
    extraer_certificado_deposito,
    procesar_pdf,
)

PDF = Path(
    r"d:\Lucas\Cursor Programacion\zAnalisis Integral del Contribuyente"
    r"\Proceso liquidaciones\Para analizar\BOCA DEL RIO S.A"
    r"\Certificados de depósito\332021153866.pdf"
)
SALIDA = PDF.parent / "332021153866_analisis"


def _buscar(texto: str, patron: str, *, flags: int = re.I) -> str | None:
    m = re.search(patron, texto, flags)
    return m.group(1).strip() if m else None


def _parse_num(val: str | None) -> float | None:
    return _parse_numero(val)


def extraer_json_completo(texto: str, paginas: list[str]) -> dict[str, Any]:
    """Estructura JSON de ingeniería inversa (todas las secciones del PDF)."""
    m_tipo = re.search(
        r"Tipo de certificado:\s*(.+?)\s+Campa[nñ]a:\s*\S+",
        texto,
        re.I | re.S,
    )
    cabecera = {
        "titulo": "CERTIFICACIÓN ELECTRÓNICA DE GRANOS",
        "fecha_emision": _buscar(texto, r"Fecha Emisi[oó]n:\s*([\d/.]+)"),
        "tipo_certificado": m_tipo.group(1).strip() if m_tipo else None,
        "campana": _buscar(texto, r"Campa[nñ]a:\s*(\S+)"),
        "grano_y_tipo_completo": _buscar(texto, r"Grano y Tipo:\s*([^\n]+)"),
        "coe": _buscar(texto, r"C\.O\.E\.:\s*(\d+)"),
    }

    m_rs = re.search(
        r"Raz[oó]n Social:\s*([^\n]+?)\s+Raz[oó]n Social:\s*([^\n]+)",
        texto,
        re.I,
    )
    depositario = {}
    depositante = {}
    if m_rs:
        depositario["razon_social"] = m_rs.group(1).strip()
        depositante["razon_social"] = m_rs.group(2).strip()
    m_cuit = re.search(r"C\.U\.I\.T\.:\s*(\d+)\s+C\.U\.I\.T\.:\s*(\d+)", texto, re.I)
    if m_cuit:
        depositario["cuit"] = m_cuit.group(1)
        depositante["cuit"] = m_cuit.group(2)

    def bloque_parte_lateral(etiqueta: str, lado: int) -> dict[str, Any]:
        """lado 1=depositario, 2=depositante (columnas en la misma línea)."""
        campos = {
            "domicilio": r"Domicilio:\s*([^\n]+?)\s+Domicilio:\s*([^\n]+)",
            "localidad": r"Localidad:\s*([^\n]+?)\s+Localidad:\s*([^\n]+)",
            "provincia": r"Provincia:\s*([^\n]+?)\s+Provincia:\s*([^\n]+)",
            "iva": r"I\.V\.A\.:\s*(\S+)\s+I\.V\.A\.:\s*(\S+)",
            "ingresos_brutos": r"Ingresos Brutos N[ºo]:\s*(\S+)\s+Ingresos Brutos N[ºo]:\s*(\S+)",
        }
        out: dict[str, Any] = {}
        for nombre, patron in campos.items():
            m = re.search(patron, texto, re.I)
            if m:
                out[nombre] = m.group(lado).strip()
        return out

    if depositario:
        depositario.update(bloque_parte_lateral("DEPOSITARIO", 1))
    if depositante:
        depositante.update(bloque_parte_lateral("DEPOSITANTE", 2))

    corredor = {}
    m_corr = re.search(
        r"CORREDOR:\s*C\.U\.I\.T\.:\s*(\d+)\s+Raz[oó]n Social:\s*([^\n]+)",
        texto,
        re.I,
    )
    if m_corr:
        corredor = {"cuit": m_corr.group(1), "razon_social": m_corr.group(2).strip()}

    tarifas = {
        "almacenaje": _parse_num(_buscar(texto, r"Almacenaje:\s*\$\s*([\d.,]+)")),
        "acarreo": _parse_num(_buscar(texto, r"Acarreo:\s*\$\s*([\d.,]+)")),
        "gastos_generales": _parse_num(
            _buscar(texto, r"Gastos Generales:\s*\$\s*([\d.,]+)")
        ),
        "zarandeo_tarifa": _parse_num(_buscar(texto, r"Zarandeo:\s*\$\s*([\d.,]+)")),
        "secado_rango": _buscar(texto, r"Secado:\s*(De [^\n]+?)Monto Secado:"),
        "monto_secado_tarifa": _parse_num(
            _buscar(texto, r"Monto Secado:\s*\$\s*([\d.,]+)")
        ),
        "por_punto_exceso": _parse_num(
            _buscar(texto, r"Por c/pto\. de exceso:\s*\$\s*([\d.,]+)")
        ),
        "otros": _parse_num(_buscar(texto, r"Otros:\s*\$\s*([\d.,]+)")),
    }

    calidad = {
        "analisis_muestra_n": _buscar(texto, r"An[aá]lisis de muestra N[ºo]:\s*(\d+)"),
        "boletin_n": _buscar(texto, r"Bolet[ií]n N[ºo]:\s*(\d+)"),
        "grado": _buscar(texto, r"^(G\d+)\s+\d+\s+\d+", flags=re.M),
        "cont_proteico": _buscar(texto, r"^G\d+\s+(\d+)\s+\d+", flags=re.M),
        "factor": _buscar(texto, r"^G\d+\s+\d+\s+(\d+)", flags=re.M),
        "rubros": [],
    }
    for rubro in ("Proteína", "Peso Hectolítrico", "Total Dañados"):
        m_r = re.search(
            rf"{rubro}\s+([\d.,]+)\s*%\s+\$\s*([\d.,]+)",
            texto,
            re.I,
        )
        if m_r:
            calidad["rubros"].append(
                {
                    "rubro": rubro,
                    "porcentaje": _parse_num(m_r.group(1)),
                    "monto_bonificacion": _parse_num(m_r.group(2)),
                }
            )

    m_grano = re.search(
        r"(\d+)\s*/\s*(\d+)\s+(\d{2}/\d{2}/\d{4})\s+([\d.,]+)\s+([\d.,]+)\s+\$\s*([\d.,]+)\s+\$\s*([\d.,]+)\s+([\d.,]+)\s+([\d.,]+)\s+\$\s*([\d.,]+)\s+\$\s*([\d.,]+)",
        texto,
    )
    linea_grano = {}
    if m_grano:
        linea_grano = {
            "ctg": m_grano.group(1),
            "carta_porte": m_grano.group(2),
            "fecha_ctg": m_grano.group(3),
            "kgs_conforme_definitivo": _parse_num(m_grano.group(4)),
            "zarandeo_merma_kgs": _parse_num(m_grano.group(5)),
            "zarandeo_tarifa": _parse_num(m_grano.group(6)),
            "zarandeo_importe": _parse_num(m_grano.group(7)),
            "secado_humedad_pct": _parse_num(m_grano.group(8)),
            "secado_merma_kgs": _parse_num(m_grano.group(9)),
            "secado_tarifa": _parse_num(m_grano.group(10)),
            "secado_importe": _parse_num(m_grano.group(11)),
        }

    idx = texto.upper().find("PESO")
    bloque_peso = texto[idx : idx + 800] if idx >= 0 else ""
    peso_servicios = {
        "kilos_forma_pago": _parse_num(
            _buscar(bloque_peso, r"^([\d.,]+)\s+Al[ií]cuota", flags=re.M)
        ),
        "alicuota_iva_pct": _buscar(bloque_peso, r"Al[ií]cuota de IVA:\s*([\d.,]+)\s*%"),
        "peso_bruto": _parse_num(_buscar(bloque_peso, r"^([\d.,]+)\s+Zarandeo:", flags=re.M)),
        "volatil": _parse_num(_buscar(bloque_peso, r"Vol[aá]til:\s*([\d.,]+)")),
        "secado_peso": _parse_num(
            _buscar(bloque_peso, r"Secado:\s*([\d.,]+)\s+Secado:", flags=re.I)
        ),
        "zarandeo_peso": _parse_num(
            _buscar(bloque_peso, r"Zarandeo:\s*([\d.,]+)\s+Otros:", flags=re.I)
        ),
        "peso_neto": _parse_num(_buscar(bloque_peso, r"Peso Neto:\s*([\d.,]+)")),
        "servicios": {
            "gastos_generales": _parse_num(
                _buscar(bloque_peso, r"Gastos Generales:\s*\$\s*([\d.,]+)")
            ),
            "importe_iva": _parse_num(
                _buscar(bloque_peso, r"Importe IVA:\s*\$\s*([\d.,]+)")
            ),
            "zarandeo": _parse_num(_buscar(bloque_peso, r"Zarandeo:\s*\$\s*([\d.,]+)")),
            "secado": _parse_num(_buscar(bloque_peso, r"Secado:\s*\$\s*([\d.,]+)")),
            "otros": _parse_num(_buscar(bloque_peso, r"Otros:\s*\$\s*([\d.,]+)")),
            "cptos_no_gravados": _parse_num(
                _buscar(bloque_peso, r"Cptos\.\s*No Gravados\s*\$?\s*([\d.,]+)")
            ),
            "percepciones_iva": _parse_num(
                _buscar(bloque_peso, r"Percepciones IVA:\s*\$\s*([\d.,]+)")
            ),
            "otras_percepciones": _parse_num(
                _buscar(bloque_peso, r"Otras percepciones\s*\$?\s*([\d.,]+)")
            ),
            "total": _parse_num(_buscar(bloque_peso, r"TOTAL:\s*\$\s*([\d.,]+)")),
        },
    }

    pag2 = paginas[1] if len(paginas) > 1 else ""
    datos_adicionales = _buscar(pag2 or texto, r"Datos Adicionales:\s*(.+?)(?:\nFirma|$)", flags=re.S)

    return {
        "coe": cabecera.get("coe"),
        "cabecera": cabecera,
        "planta_nro": _buscar(texto, r"PLANTA NRO:\s*(\d+)"),
        "depositario": depositario,
        "depositante": depositante,
        "corredor": corredor,
        "tarifas_cada_100_kgrs": tarifas,
        "calidad": calidad,
        "linea_granos": linea_grano,
        "peso_y_servicios": peso_servicios,
        "pagina_2": {
            "datos_adicionales": (datos_adicionales or "").strip(),
        },
        "extraccion_excel_actual": extraer_certificado_deposito(texto),
    }


def _aplanar(d: dict[str, Any], prefijo: str = "") -> dict[str, Any]:
    out: dict[str, Any] = {}
    for k, v in d.items():
        clave = f"{prefijo}{k}" if not prefijo else f"{prefijo}.{k}"
        if isinstance(v, dict):
            out.update(_aplanar(v, clave))
        elif isinstance(v, list):
            out[clave] = json.dumps(v, ensure_ascii=False)
        else:
            out[clave] = v
    return out


def escribir_excel_analisis(ruta: Path, json_full: dict[str, Any]) -> None:
    wb = Workbook()

    ws1 = wb.active
    ws1.title = "Formato actual VL"
    fila_actual = json_full["extraccion_excel_actual"]
    ws1.append(list(_COLS_CERTIFICADOS))
    for c in ws1[1]:
        c.font = Font(bold=True)
    from liquidaciones_pdf import _fila_certificado_a_lista

    ws1.append(_fila_certificado_a_lista(fila_actual))

    ws2 = wb.create_sheet("JSON completo")
    plano = _aplanar({k: v for k, v in json_full.items() if k != "extraccion_excel_actual"})
    ws2.append(["Campo", "Valor"])
    for c in ws2[1]:
        c.font = Font(bold=True)
    for k in sorted(plano.keys()):
        ws2.append([k, plano[k]])

    ws3 = wb.create_sheet("Comparacion sistemas")
    ws3.append(["Sistema", "Ventajas", "Limitaciones", "Recomendacion"])
    for c in ws3[1]:
        c.font = Font(bold=True)
    filas_cmp = [
        (
            "pdfplumber + regex (actual)",
            "Ya integrado en liquidaciones_pdf.py; funciona con este PDF; bajo costo",
            "Depende del orden del texto; tablas complejas pueden variar entre PDF",
            "MANTENER como base para el Excel resumen VL",
        ),
        (
            "pdfplumber extract_tables()",
            "Estructura tabular explícita (depositario/depositante separados)",
            "Celdas fusionadas, nulls y filas rotas; difícil de generalizar",
            "Complemento puntual, no motor principal",
        ),
        (
            "JSON completo (ingeniería inversa)",
            "Captura CTG, calidad, tarifas, corredor, nº interno",
            "Más campos de los que pide el Excel resumen hoy",
            "Usar para ampliar plantilla si el negocio lo requiere",
        ),
    ]
    for row in filas_cmp:
        ws3.append(list(row))

    wb.save(ruta)


def main() -> None:
    SALIDA.mkdir(parents=True, exist_ok=True)
    texto, paginas = _leer_texto_pdf(PDF)
    json_full = extraer_json_completo(texto, paginas)
    tipo, _ = procesar_pdf(PDF)

    payload = {
        "archivo": PDF.name,
        "tipo_detectado": tipo,
        "paginas": len(paginas),
        "ingenieria_inversa": json_full,
    }
    json_path = SALIDA / "332021153866.json"
    json_path.write_text(
        json.dumps(payload, indent=2, ensure_ascii=False, default=str),
        encoding="utf-8",
    )

    excel_resumen = SALIDA / "332021153866_resumen_VL.xlsx"
    escribir_excel_resumen(
        excel_resumen,
        certificados=[json_full["extraccion_excel_actual"]],
        lpg=[],
        hacienda=[],
    )

    excel_analisis = SALIDA / "332021153866_analisis.xlsx"
    escribir_excel_analisis(excel_analisis, json_full)

    print("JSON:", json_path)
    print("Excel resumen VL:", excel_resumen)
    print("Excel analisis:", excel_analisis)


if __name__ == "__main__":
    main()
