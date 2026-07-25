"""Extracción de datos desde PDF de liquidaciones (certificado, LPG, hacienda) a Excel."""

from __future__ import annotations

import io
import re
import unicodedata
from datetime import date, datetime
from pathlib import Path
from typing import Any

import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils.dataframe import dataframe_to_rows

SHEET_CERTIFICADOS = "Certificados deposito"
SHEET_LPG = "Liq Primaria Granos"
SHEET_HACIENDA = "Liquidacion Hacienda"
NOMBRE_EXCEL_RESUMEN = "Resumen liquidaciones.xlsx"

MSG_AJUSTE_UNIFICADO = "Verificar porque es ajuste unificado"

_COLS_CERTIFICADOS = [
    "Fecha emision",
    "COE",
    "Campaña",
    "Tipo certificado",
    "Grano y tipo",
    "Depositario Razon Social",
    "Depositario CUIT",
    "Depositante Razon social",
    "Depositante CUIT",
    "Peso Bruto",
    "Volatil",
    "Secado",
    "Zarandeo",
    "Peso Neto",
    "Zarandeo",
    "Secado",
    "Otros",
    "Gastos Generales",
    "Importe IVA",
    "Cptos. No Gravados",
    "Percepciones IVA",
    "Otras Pecepciones",
    "Total",
    "CUIT Receptor",
    "Razon Social Receptor",
]

_COLS_LPG = [
    "Fecha",
    "COE",
    "Razón Social Comprador",
    "CUIT Comprador",
    "COES Relacionados",
    "Precio/TN",
    "Grado",
    "Grano",
    "Flete por TN",
    "Puerto",
    "Cantidad",
    "Precio/Kg",
    "Subtotal",
    "% Alicuota IVA",
    "Importe IVA",
    "Operación c/IVA",
    "Deducciones Bruto",
    "Deducciones IVA",
    "Retenciones IVA",
    "Retenciones IG",
    "Otras Retenciones",
]

_COLS_HACIENDA = [
    "Fecha",
    "N° Liquidacion",
    "Nombre Receptor",
    "CUIT Receptor",
    "N° DTE",
    "N° Renspa",
    "Importe Bruto",
    "IVA s/Bruto",
    "Total Gastos",
    "IVA s/Gastos",
    "Total tributos",
    "Importe Neto",
    "Categoria / Raza",
    "UM",
    "Cantidad",
    "$ UM",
    "$ Bruto",
    "% IVA",
    "$ IVA",
]


def _norm(text: object) -> str:
    s = unicodedata.normalize("NFKD", str(text or ""))
    s = "".join(c for c in s if not unicodedata.combining(c))
    return s


def _norm_lower(text: object) -> str:
    return _norm(text).lower()


def _solo_digitos(val: object) -> str:
    return re.sub(r"\D", "", str(val or ""))


def _parse_numero(val: object) -> float | None:
    if val is None:
        return None
    s = str(val).strip()
    if not s or s.lower() in ("-", "—"):
        return None
    s = re.sub(r"^\$+\s*", "", s)
    s = s.replace(" ", "")
    if not s:
        return None
    if "," in s and "." in s:
        if s.rfind(",") > s.rfind("."):
            s = s.replace(".", "").replace(",", ".")
        else:
            s = s.replace(",", "")
    elif "," in s:
        s = s.replace(".", "").replace(",", ".")
    else:
        s = s.replace(",", "")
    try:
        return float(s)
    except ValueError:
        return None


def _parse_fecha(val: object) -> date | None:
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    s = str(val).strip()
    m = re.search(r"(\d{1,2})[/.-](\d{1,2})[/.-](\d{2,4})", s)
    if not m:
        return None
    d, mo, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
    if y < 100:
        y += 2000
    try:
        return date(y, mo, d)
    except ValueError:
        return None


def _buscar(texto: str, patron: str, *, flags: int = re.I) -> str | None:
    m = re.search(patron, texto, flags)
    return m.group(1).strip() if m else None


def _buscar_num(texto: str, patron: str) -> float | None:
    val = _buscar(texto, patron)
    return _parse_numero(val) if val is not None else None


def _leer_texto_pdf(ruta: Path) -> tuple[str, list[str]]:
    paginas: list[str] = []
    with pdfplumber.open(ruta) as pdf:
        for page in pdf.pages:
            paginas.append(page.extract_text() or "")
    return "\n".join(paginas), paginas


def _clasificar_pdf(texto: str, paginas: list[str]) -> str:
    t = _norm_lower(texto[:2000])
    if "certificacion electronica de granos" in t or "certificaci" in t and "granos" in t[:400]:
        if "tipo de certificado" in t:
            return "certificado"
    if "liquidacion primaria de granos" in t or "liquidacion secundaria de granos" in t:
        return "lpg"
    if "liquidacion de venta directa" in t or (
        "liquidaci" in t and "hacienda" not in t and "granos" not in t[:500]
        and re.search(r"n[°o]\s*\d{5}-\d{8}", t)
    ):
        return "hacienda"
    p0 = _norm_lower(paginas[0] if paginas else "")
    if "original" in p0 and "liquidaci" in p0:
        return "hacienda"
    return "desconocido"


def _extraer_bloque_peso_servicios(texto: str) -> tuple[dict[str, float | None], dict[str, float | None]]:
    peso: dict[str, float | None] = {
        "Peso Bruto": None,
        "Volatil": None,
        "Secado": None,
        "Zarandeo": None,
        "Peso Neto": None,
    }
    servicios: dict[str, float | None] = {
        "Zarandeo": None,
        "Secado": None,
        "Otros": None,
        "Gastos Generales": None,
        "Importe IVA": None,
        "Cptos. No Gravados": None,
        "Percepciones IVA": None,
        "Otras Pecepciones": None,
        "Total": None,
    }
    idx = texto.upper().find("PESO")
    if idx < 0:
        return peso, servicios
    bloque = texto[idx : idx + 1200]

    m_bruto = re.search(r"^([\d.,]+)\s+Al[ií]cuota", bloque, re.M | re.I)
    if m_bruto:
        peso["Peso Bruto"] = _parse_numero(m_bruto.group(1))

    m_vol = re.search(r"^([\d.,]+)\s+Zarandeo:", bloque, re.M | re.I)
    if m_vol:
        peso["Volatil"] = _parse_numero(m_vol.group(1))

    m_sec_peso = re.search(r"Secado:\s*([\d.,]+)\s+Secado:", bloque, re.I)
    if m_sec_peso:
        peso["Secado"] = _parse_numero(m_sec_peso.group(1))

    m_zar_peso = re.search(r"Zarandeo:\s*([\d.,]+)\s+Otros:", bloque, re.I)
    if m_zar_peso:
        peso["Zarandeo"] = _parse_numero(m_zar_peso.group(1))

    m_neto = re.search(r"Peso Neto:\s*([\d.,]+)", bloque, re.I)
    if m_neto:
        peso["Peso Neto"] = _parse_numero(m_neto.group(1))

    servicios["Gastos Generales"] = _buscar_num(bloque, r"Gastos Generales:\s*\$?\s*([\d.,]+)")
    servicios["Importe IVA"] = _buscar_num(bloque, r"Importe IVA:\s*\$?\s*([\d.,]+)")
    servicios["Cptos. No Gravados"] = _buscar_num(
        bloque, r"Cptos\.\s*No Gravados\s*\$?\s*([\d.,]+)"
    )
    servicios["Percepciones IVA"] = _buscar_num(
        bloque, r"Percepciones IVA:\s*\$?\s*([\d.,]+)"
    )
    servicios["Otras Pecepciones"] = _buscar_num(
        bloque, r"Otras percepciones\s*\$?\s*([\d.,]+)"
    )
    servicios["Total"] = _buscar_num(bloque, r"TOTAL:\s*\$?\s*([\d.,]+)")
    servicios["Zarandeo"] = _buscar_num(bloque, r"Zarandeo:\s*\$\s*([\d.,]+)")
    servicios["Secado"] = _buscar_num(bloque, r"Secado:\s*\$\s*([\d.,]+)")
    servicios["Otros"] = _buscar_num(bloque, r"Otros:\s*\$\s*([\d.,]+)")
    return peso, servicios


_CAMPOS_NUMERICOS_CERTIFICADO = (
    "Peso Bruto",
    "Volatil",
    "Secado",
    "Zarandeo",
    "Peso Neto",
    "Zarandeo_serv",
    "Secado_serv",
    "Otros",
    "Gastos Generales",
    "Importe IVA",
    "Cptos. No Gravados",
    "Percepciones IVA",
    "Otras Pecepciones",
    "Total",
)


def _tipo_certificado_base(tipo: object) -> str:
    if not tipo:
        return ""
    parte = str(tipo).split("-", 1)[0].strip()
    return _norm_lower(parte)


def _es_certificado_anulado(texto: str) -> bool:
    muestra = _norm_lower(texto[:2500])
    return bool(
        re.search(r"comprobante\s+anulado|certificado\s+anulado", muestra, re.I)
    )


def _extraer_receptor_certificado(texto: str) -> tuple[str | None, str | None]:
    m = re.search(
        r"RECEPTOR:\s*C\.U\.I\.T\.:\s*(\d+)\s+Raz[oó]n Social:\s*([^\n]+)",
        texto,
        re.I,
    )
    if m:
        return m.group(1), m.group(2).strip()
    m = re.search(
        r"RECEPTOR.*?C\.U\.I\.T\.:\s*(\d+).*?Raz[oó]n Social:\s*([^\n]+)",
        texto,
        re.I | re.S,
    )
    if m:
        return m.group(1), m.group(2).strip()
    return None, None


def _extraer_kilos_retiro(texto: str) -> float | None:
    idx = texto.upper().find("DETALLE DE RETIRO")
    bloque = texto[idx : idx + 900] if idx >= 0 else texto
    m = re.search(r"Kilos:\s*([\d.,]+)", bloque, re.I)
    if m:
        return _parse_numero(m.group(1))
    m = re.search(r"([\d.,]+)\s*Kg\b", bloque, re.I)
    return _parse_numero(m.group(1)) if m else None


def _vaciar_numericos_certificado(fila: dict[str, Any]) -> None:
    for campo in _CAMPOS_NUMERICOS_CERTIFICADO:
        fila[campo] = None


def _aplicar_reglas_tipo_certificado(fila: dict[str, Any], texto: str) -> dict[str, Any]:
    fila["CUIT Receptor"] = None
    fila["Razon Social Receptor"] = None

    if _es_certificado_anulado(texto):
        _vaciar_numericos_certificado(fila)
        return fila

    tipo_base = _tipo_certificado_base(fila.get("Tipo certificado"))
    if tipo_base == "transferencia":
        cuit, razon = _extraer_receptor_certificado(texto)
        fila["CUIT Receptor"] = int(cuit) if cuit else None
        fila["Razon Social Receptor"] = razon
        return fila

    if tipo_base == "retiro":
        _vaciar_numericos_certificado(fila)
        kilos = _extraer_kilos_retiro(texto)
        if kilos is not None:
            fila["Peso Bruto"] = -abs(kilos)
        return fila

    return fila


def _extraer_grano_y_tipo(texto: str) -> str | None:
    m = re.search(r"Grano y Tipo:\s*(.+)", texto, re.I)
    if not m:
        return None
    raw = re.split(r"\s+C\.O\.E\.:", m.group(1), maxsplit=1, flags=re.I)[0].strip()
    raw = re.sub(r"\s*-\s*$", "", raw).strip()
    if " - " in raw:
        raw = raw.split(" - ", 1)[0].strip()
    return raw or None


def extraer_certificado_deposito(texto: str) -> dict[str, Any]:
    fecha = _parse_fecha(_buscar(texto, r"Fecha Emisi[oó]n:\s*([\d/.]+)"))
    coe = _solo_digitos(_buscar(texto, r"C\.O\.E\.:\s*(\d+)") or "")
    grano = _extraer_grano_y_tipo(texto)

    m_rs = re.search(
        r"Raz[oó]n Social:\s*([^\n]+?)\s+Raz[oó]n Social:\s*([^\n]+)",
        texto,
        re.I,
    )
    dep_rs = m_rs.group(1).strip() if m_rs else None
    cte_rs = m_rs.group(2).strip() if m_rs else None
    m_cuit = re.search(r"C\.U\.I\.T\.:\s*(\d+)\s+C\.U\.I\.T\.:\s*(\d+)", texto, re.I)
    dep_cuit = m_cuit.group(1) if m_cuit else ""
    cte_cuit = m_cuit.group(2) if m_cuit else ""

    m_tipo_camp = re.search(
        r"Tipo de certificado:\s*(.+?)\s+Campa[nñ]a:\s*(\S+)",
        texto,
        re.I | re.S,
    )
    tipo_certificado = m_tipo_camp.group(1).strip() if m_tipo_camp else None
    campana = m_tipo_camp.group(2).strip() if m_tipo_camp else None

    peso, servicios = _extraer_bloque_peso_servicios(texto)

    fila: dict[str, Any] = {
        "Fecha emision": fecha,
        "COE": int(coe) if coe else None,
        "Campaña": campana,
        "Tipo certificado": tipo_certificado,
        "Grano y tipo": grano,
        "Depositario Razon Social": dep_rs,
        "Depositario CUIT": int(dep_cuit) if dep_cuit else None,
        "Depositante Razon social": cte_rs,
        "Depositante CUIT": int(cte_cuit) if cte_cuit else None,
        "Peso Bruto": peso["Peso Bruto"],
        "Volatil": peso["Volatil"],
        "Secado": peso["Secado"],
        "Zarandeo": peso["Zarandeo"],
        "Peso Neto": peso["Peso Neto"],
        "Zarandeo_serv": servicios["Zarandeo"],
        "Secado_serv": servicios["Secado"],
        "Otros": servicios["Otros"],
        "Gastos Generales": servicios["Gastos Generales"],
        "Importe IVA": servicios["Importe IVA"],
        "Cptos. No Gravados": servicios["Cptos. No Gravados"],
        "Percepciones IVA": servicios["Percepciones IVA"],
        "Otras Pecepciones": servicios["Otras Pecepciones"],
        "Total": servicios["Total"],
    }
    return _map_certificado_a_columnas(_aplicar_reglas_tipo_certificado(fila, texto))


def _map_certificado_a_columnas(fila: dict[str, Any]) -> dict[str, Any]:
    """Dos columnas Zarandeo/Secado: peso y servicios."""
    return {
        "Fecha emision": fila.get("Fecha emision"),
        "COE": fila.get("COE"),
        "Campaña": fila.get("Campaña"),
        "Tipo certificado": fila.get("Tipo certificado"),
        "Grano y tipo": fila.get("Grano y tipo"),
        "Depositario Razon Social": fila.get("Depositario Razon Social"),
        "Depositario CUIT": fila.get("Depositario CUIT"),
        "Depositante Razon social": fila.get("Depositante Razon social"),
        "Depositante CUIT": fila.get("Depositante CUIT"),
        "Peso Bruto": fila.get("Peso Bruto"),
        "Volatil": fila.get("Volatil"),
        "Secado": fila.get("Secado"),
        "Zarandeo": fila.get("Zarandeo"),
        "Peso Neto": fila.get("Peso Neto"),
        "Zarandeo_serv": fila.get("Zarandeo_serv"),
        "Secado_serv": fila.get("Secado_serv"),
        "Otros": fila.get("Otros"),
        "Gastos Generales": fila.get("Gastos Generales"),
        "Importe IVA": fila.get("Importe IVA"),
        "Cptos. No Gravados": fila.get("Cptos. No Gravados"),
        "Percepciones IVA": fila.get("Percepciones IVA"),
        "Otras Pecepciones": fila.get("Otras Pecepciones"),
        "Total": fila.get("Total"),
        "CUIT Receptor": fila.get("CUIT Receptor"),
        "Razon Social Receptor": fila.get("Razon Social Receptor"),
    }


def _parse_lpg_condiciones(texto_pag: str) -> tuple[float | None, str | None, str | None, float | None, str | None]:
    """Devuelve precio_tn, grado, grano, flete, puerto."""
    m_aj = re.search(
        r"Precio/TN Grado Factor Grano Flete por TN Puerto\s*\n"
        r"\$?\s*([\d.,]+)\s+\S*\s+(.+?)\s+\$\s*([\d.,]+)\s+(\S+)",
        texto_pag,
        re.I | re.S,
    )
    if m_aj:
        return (
            _parse_numero(m_aj.group(1)),
            None,
            m_aj.group(2).strip(),
            _parse_numero(m_aj.group(3)),
            m_aj.group(4).strip(),
        )
    m = re.search(
        r"Precio/TN Grado Grano Flete por TN Puerto\s*\n"
        r"\$\s*([\d.,]+)\s+(\S*)\s+(.+?)\s+\$\s*([\d.,]+)\s+(\S+)",
        texto_pag,
        re.I | re.S,
    )
    if not m:
        return None, None, None, None, None
    precio = _parse_numero(m.group(1))
    tok_grado = m.group(2).strip()
    resto = m.group(3).strip()
    flete = _parse_numero(m.group(4))
    puerto = m.group(5).strip()
    if tok_grado and re.fullmatch(r"\d+", tok_grado):
        grano = f"{tok_grado} {resto}".strip()
        grado = None
    elif not tok_grado and re.match(r"\d+\s*-", resto):
        grano = resto
        grado = None
    else:
        grado = tok_grado or None
        grano = resto
    return precio, grado, grano, flete, puerto


def _cuit_comprador_lpg(texto_pag: str) -> str:
    m = re.search(
        r"COMPRADOR\s+VENDEDOR.*?C\.U\.I\.T\.:\s*(\d+).*?C\.U\.I\.T\.:\s*(\d+)",
        texto_pag,
        re.I | re.S,
    )
    return m.group(1) if m else _solo_digitos(_buscar(texto_pag, r"C\.U\.I\.T\.:\s*(\d+)") or "")


def _sumar_importes_deducciones(texto: str) -> tuple[float, float]:
    bruto = 0.0
    iva = 0.0
    idx = texto.upper().find("DEDUCCIONES")
    if idx < 0:
        return 0.0, 0.0
    bloque = texto[idx : idx + 2500]
    if "RETENCIONES" in bloque.upper():
        bloque = bloque.split("RETENCIONES")[0]
    for linea in bloque.splitlines():
        ln = linea.strip()
        lnl = _norm_lower(ln)
        if not lnl.startswith(("otras deducc", "comisi")):
            continue
        montos = [_parse_numero(x) for x in re.findall(r"\$\s*([\d.,]+)", ln)]
        montos = [m for m in montos if m is not None]
        if not montos:
            continue
        bruto += montos[0]
        if len(montos) >= 3:
            iva += montos[-2]
    return round(bruto, 2), round(iva, 2)


def _sumar_retenciones(texto: str) -> tuple[float, float, float]:
    idx = texto.upper().find("RETENCIONES")
    if idx < 0:
        return 0.0, 0.0, 0.0
    bloque = texto[idx : idx + 2000]
    if "IMPORTES TOTALES" in bloque.upper():
        bloque = bloque.split("IMPORTES TOTALES")[0]
    iva = 0.0
    ig = 0.0
    otras = 0.0
    for linea in bloque.splitlines():
        ln = _norm_lower(linea)
        if "retenc" not in ln and "ret." not in ln:
            continue
        nums = [_parse_numero(x) for x in re.findall(r"\$\s*([\d.,]+)", linea)]
        nums = [n for n in nums if n is not None]
        if not nums:
            continue
        ret = nums[-1]
        if "gananc" in ln:
            ig += ret
        elif "i.v.a" in ln or ("iva" in ln and "gananc" not in ln):
            iva += ret
        else:
            otras += ret
    return round(iva, 2), round(ig, 2), round(otras, 2)


def extraer_lpg(texto: str, paginas: list[str]) -> dict[str, Any]:
    texto_full = texto
    if "AJUSTE UNIFICADO" in texto_full.upper():
        pagina_datos = paginas[0] if paginas else texto
        fecha_hdr = _parse_fecha(_buscar(pagina_datos, r"^(\d{2}/\d{2}/\d{4})", flags=re.M))
        coe = _solo_digitos(_buscar(texto_full, r"C\.O\.E\.:\s*(\d+)") or "")
        comprador = _buscar(
            pagina_datos,
            r"COMPRADOR\s+VENDEDOR\s+Raz[oó]n Social:\s*([^\n]+?)\s+Raz[oó]n Social:",
        )
        cuit_comprador = _cuit_comprador_lpg(pagina_datos)
        coes_rel: list[str] = []
        m_coes = re.search(r"COES RELACIONADOS\s*(.*?)\s*MERCADERIA", pagina_datos, re.S | re.I)
        if m_coes:
            coes_rel = re.findall(r"\b(\d{10,15})\b", m_coes.group(1))

        precio_tn = grano = flete = puerto = None
        for pag in paginas:
            pt, _g, gr, fl, pt_p = _parse_lpg_condiciones(pag)
            if pt is not None:
                precio_tn, grano, flete, puerto = pt, gr, fl, pt_p
                break

        return {
            "Fecha": fecha_hdr,
            "COE": int(coe) if coe else None,
            "Razón Social Comprador": comprador,
            "CUIT Comprador": int(cuit_comprador) if cuit_comprador else None,
            "COES Relacionados": " / ".join(coes_rel) if coes_rel else None,
            "Precio/TN": precio_tn,
            "Grado": None,
            "Grano": grano,
            "Flete por TN": flete,
            "Puerto": puerto,
            "Cantidad": MSG_AJUSTE_UNIFICADO,
            "Precio/Kg": None,
            "Subtotal": None,
            "% Alicuota IVA": None,
            "Importe IVA": None,
            "Operación c/IVA": None,
            "Deducciones Bruto": None,
            "Deducciones IVA": None,
            "Retenciones IVA": None,
            "Retenciones IG": None,
            "Otras Retenciones": None,
        }

    pagina = paginas[0] if paginas else texto
    fecha = _parse_fecha(_buscar(pagina, r"^(\d{2}/\d{2}/\d{4})", flags=re.M))
    coe = _solo_digitos(_buscar(texto_full, r"C\.O\.E\.:\s*(\d+)") or "")
    comprador = _buscar(
        pagina,
        r"COMPRADOR\s+VENDEDOR\s+Raz[oó]n Social:\s*([^\n]+?)\s+Raz[oó]n Social:",
    )
    cuit_comprador = _cuit_comprador_lpg(pagina)

    coes_rel: list[str] = []
    m_coes = re.search(r"COES RELACIONADOS\s*(.*?)\s*MERCADERIA", pagina, re.S | re.I)
    if m_coes:
        coes_rel = re.findall(r"\b(\d{10,15})\b", m_coes.group(1))

    precio_tn, grado, grano, flete, puerto = _parse_lpg_condiciones(pagina)

    cantidad = precio_kg = subtotal = alicuota = imp_iva = op_iva = None
    m_op = re.search(
        r"OPERACI[^\n]*\nCantidad Precio/Kg Subtotal.*?Operaci[^\n]*\n"
        r"([\d.,]+)\s+Kg\s+\$?([\d.,]+)\s+\$?([\d.,]+)\s+([\d.,]+)\s+\$?([\d.,]+)\s+\$?([\d.,]+)",
        pagina,
        re.I | re.S,
    )
    if m_op:
        cantidad = _parse_numero(m_op.group(1))
        precio_kg = _parse_numero(m_op.group(2))
        subtotal = _parse_numero(m_op.group(3))
        alicuota = _parse_numero(m_op.group(4))
        imp_iva = _parse_numero(m_op.group(5))
        op_iva = _parse_numero(m_op.group(6))

    ded_bruto, ded_iva = _sumar_importes_deducciones(pagina)
    ret_iva, ret_ig, ret_otras = _sumar_retenciones(pagina)

    return {
        "Fecha": fecha,
        "COE": int(coe) if coe else None,
        "Razón Social Comprador": comprador,
        "CUIT Comprador": int(cuit_comprador) if cuit_comprador else None,
        "COES Relacionados": " / ".join(coes_rel) if coes_rel else None,
        "Precio/TN": precio_tn,
        "Grado": grado,
        "Grano": grano,
        "Flete por TN": flete,
        "Puerto": puerto,
        "Cantidad": cantidad,
        "Precio/Kg": precio_kg,
        "Subtotal": subtotal,
        "% Alicuota IVA": alicuota,
        "Importe IVA": imp_iva,
        "Operación c/IVA": op_iva,
        "Deducciones Bruto": ded_bruto,
        "Deducciones IVA": ded_iva,
        "Retenciones IVA": ret_iva,
        "Retenciones IG": ret_ig,
        "Otras Retenciones": ret_otras,
    }


def _paginas_hacienda(paginas: list[str]) -> list[int]:
    n = len(paginas)
    if n <= 1:
        return [0]
    if n == 3:
        return [0]
    if n == 6:
        return [0, 1]
    if n % 3 == 0:
        return list(range(0, n, 3))
    return [0]


def _parsear_animales_hacienda(texto: str) -> list[dict[str, Any]]:
    m_sec = re.search(
        r"Categor[ií]a / Raza.*?Importe Bruto:",
        texto,
        re.I | re.S,
    )
    if not m_sec:
        return []
    sec = texto[m_sec.start() : m_sec.end()]
    lineas = [ln.strip() for ln in sec.splitlines() if ln.strip()]
    try:
        start = next(
            i
            for i, ln in enumerate(lineas)
            if "categoria" in _norm_lower(ln) and "raza" in _norm_lower(ln)
        )
    except StopIteration:
        return []
    lineas = lineas[start + 1 :]

    animales: list[dict[str, Any]] = []
    i = 0
    while i < len(lineas):
        ln = lineas[i]
        if re.match(r"^Importe Bruto:", ln, re.I):
            break
        if not re.match(r"^(Bovino|Porcino|Ovino|Caprino|Equino|Camel|Ciervo)", ln, re.I):
            i += 1
            continue

        cat_partes = [ln]
        j = i + 1
        um_line = None
        while j < len(lineas):
            cand = lineas[j]
            if re.match(r"^Importe Bruto:", cand, re.I):
                break
            if re.match(r"^(Cabeza|Kg|Kilo|Litros|Unidad)", cand, re.I):
                um_line = cand
                j += 1
                break
            if re.match(r"^(Bovino|Porcino|Ovino|Caprino|Equino|Camel|Ciervo)", cand, re.I):
                break
            cat_partes.append(cand)
            j += 1

        if not um_line:
            i += 1
            continue

        cat_line0 = cat_partes[0]
        um = None
        cantidad = None
        bruto = None
        pct_iva = None
        iva = None
        um_val = None

        m_um = re.match(
            r"^(Cabeza|Kg|Kilos|Litros|Unidad)\s+(\d+)\s+([\d.,]+)\s+([\d.,]+)\s+([\d.,]+)\s*$",
            um_line,
            re.I,
        )
        if m_um:
            um = m_um.group(1)
            cantidad = _parse_numero(m_um.group(2))
            bruto = _parse_numero(m_um.group(3))
            pct_iva = _parse_numero(m_um.group(4))
            iva = _parse_numero(m_um.group(5))
        else:
            m_um2 = re.match(
                r"^(Cabeza|Kg|Kilos|Litros|Unidad)\s+(\d+)\s+([\d.,]+)\s+([\d.,]+)\s*$",
                um_line,
                re.I,
            )
            if m_um2:
                um = m_um2.group(1)
                cantidad = _parse_numero(m_um2.group(2))
                bruto = _parse_numero(m_um2.group(3))
                pct_iva = _parse_numero(m_um2.group(4))

        nums_cat = [_parse_numero(x) for x in re.findall(r"([\d.,]+)", cat_line0.split("/")[-1])]
        nums_cat = [n for n in nums_cat if n is not None]
        if nums_cat and um_val is None:
            um_val = nums_cat[0]
        if len(nums_cat) >= 2 and iva is None:
            iva = nums_cat[1]

        if j < len(lineas):
            extra_ln = lineas[j]
            if not re.match(
                r"^(Bovino|Porcino|Ovino|Caprino|Equino|Camel|Ciervo|Importe)",
                extra_ln,
                re.I,
            ) and not re.match(r"^(Cabeza|Kg|Kilos|Litros|Unidad)", extra_ln, re.I):
                cat_partes.append(extra_ln)
                if iva is None:
                    tail = re.findall(r"([\d.,]+)", extra_ln)
                    if tail:
                        iva = _parse_numero(tail[-1])
                j += 1

        m_precio = re.search(r"/\s*([\d.,]+)", cat_line0)
        if m_precio:
            um_val = _parse_numero(m_precio.group(1))

        cat_nombre = cat_line0
        if m_precio:
            cat_nombre = cat_line0[: m_precio.start()].strip()
        extras = []
        for extra in cat_partes[1:]:
            m_ocr = re.search(r"(Otra\s*\([^)]+\)|[^\n]+)", extra, re.I)
            ex = m_ocr.group(1).strip() if m_ocr else extra.strip()
            ex = re.sub(r"\s+\d+(\s+\d+)?\s*$", "", ex).strip()
            if ex and not re.match(r"^[\d.,\s]+$", ex):
                extras.append(ex)
        if extras:
            cat_nombre = cat_nombre + " /\n" + "\n".join(extras)
        else:
            cat_nombre = cat_nombre.rstrip("/").strip()

        animales.append(
            {
                "Categoria / Raza": cat_nombre,
                "UM": um,
                "Cantidad": cantidad,
                " $ UM": um_val,
                "$ UM": um_val,
                "$ Bruto": bruto,
                "% IVA": pct_iva,
                "$ IVA": iva,
            }
        )
        i = j

    return animales


def extraer_hacienda_pagina(texto: str) -> tuple[dict[str, Any], list[dict[str, Any]]]:
    fecha = _parse_fecha(_buscar(texto, r"Fecha\s+(\d{2}/\d{2}/\d{4})"))
    n_liq = _buscar(texto, r"N[°ºo]\s*([\d -]+)")
    if n_liq:
        n_liq = re.sub(r"\s+", "", n_liq)

    nombre = _buscar(texto, r"Nombre y Apellido:\s*([^\n]+)")
    cuit = _solo_digitos(_buscar(texto, r"Receptor\s*\nCUIT:\s*(\d+)"))
    dte = _buscar(texto, r"DTE\s+([\d-]+)")
    renspa = _buscar(texto, r"DTE\s+[\d-]+\s+([\d./]+)")

    cabecera = {
        "Fecha": fecha,
        "N° Liquidacion": n_liq,
        "Nombre Receptor": nombre,
        "CUIT Receptor": int(cuit) if cuit else None,
        "N° DTE": dte,
        "N° Renspa": renspa,
        "Importe Bruto": _buscar_num(texto, r"Importe Bruto:\s*\$?\s*([\d.,]+)"),
        "IVA s/Bruto": _buscar_num(texto, r"IVA s/Bruto:\s*\$?\s*([\d.,]+)"),
        "Total Gastos": _buscar_num(texto, r"Total Gastos:\s*\$?\s*([\d.,-]*)"),
        "IVA s/Gastos": _buscar_num(texto, r"IVA s/Gastos:\s*\$?\s*([\d.,-]*)"),
        "Total tributos": _buscar_num(texto, r"Total Tributos:\s*\$?\s*([\d.,-]*)"),
        "Importe Neto": _buscar_num(texto, r"Importe Neto:\s*\$?\s*([\d.,]+)"),
    }
    if cabecera["Total Gastos"] is None:
        cabecera["Total Gastos"] = 0.0
    if cabecera["IVA s/Gastos"] is None:
        cabecera["IVA s/Gastos"] = 0.0
    if cabecera["Total tributos"] is None:
        cabecera["Total tributos"] = 0.0

    animales = _parsear_animales_hacienda(texto)
    return cabecera, animales


def _filas_hacienda_excel(cabecera: dict[str, Any], animales: list[dict[str, Any]]) -> list[dict[str, Any]]:
    if not animales:
        return [cabecera]
    filas: list[dict[str, Any]] = []
    for idx, ani in enumerate(animales):
        if idx == 0:
            fila = dict(cabecera)
        else:
            fila = {k: None for k in _COLS_HACIENDA}
        for k in ("Categoria / Raza", "UM", "Cantidad", "$ UM", "$ Bruto", "% IVA", "$ IVA"):
            fila[k] = ani.get(k)
        filas.append(fila)
    return filas


def procesar_pdf(ruta: Path) -> tuple[str, Any]:
    texto, paginas = _leer_texto_pdf(ruta)
    tipo = _clasificar_pdf(texto, paginas)
    if tipo == "certificado":
        return tipo, extraer_certificado_deposito(texto)
    if tipo == "lpg":
        return tipo, extraer_lpg(texto, paginas)
    if tipo == "hacienda":
        filas: list[dict[str, Any]] = []
        for idx in _paginas_hacienda(paginas):
            cab, animales = extraer_hacienda_pagina(paginas[idx])
            filas.extend(_filas_hacienda_excel(cab, animales))
        return tipo, filas
    raise ValueError(f"No se reconoció el tipo de PDF: {ruta.name}")


def listar_pdfs_carpeta_cuit(carpeta_cuit: Path | str) -> list[Path]:
    """Todos los PDF bajo la carpeta del contribuyente (subcarpetas incluidas)."""
    base = Path(carpeta_cuit)
    return sorted(p for p in base.rglob("*.pdf") if p.is_file())


def procesar_lista_pdfs(
    pdfs: list[Path],
    *,
    solo_certificados: bool = False,
) -> dict[str, list[dict[str, Any]]]:
    out: dict[str, list[dict[str, Any]]] = {
        "certificado": [],
        "lpg": [],
        "hacienda": [],
    }
    for ruta in pdfs:
        try:
            tipo, datos = procesar_pdf(ruta)
        except ValueError:
            continue
        if solo_certificados and tipo != "certificado":
            continue
        if tipo == "hacienda":
            out["hacienda"].extend(datos)
        elif tipo == "lpg":
            out["lpg"].append(datos)
        else:
            out[tipo].append(datos)
    if not solo_certificados:
        out["lpg"].sort(
            key=lambda r: (r.get("Fecha") or date.min, r.get("COE") or 0),
            reverse=True,
        )
    return out


def actualizar_excel_resumen_carpeta_cuit(carpeta_cuit: Path | str) -> Path:
    """
    Regenera el Excel resumen en la carpeta del CUIT a partir de todos los PDF presentes.
    Se invoca tras cada descarga para ir completando el archivo durante el proceso VL.
    """
    carpeta = Path(carpeta_cuit)
    pdfs = listar_pdfs_carpeta_cuit(carpeta)
    datos = procesar_lista_pdfs(pdfs, solo_certificados=True)
    destino = carpeta / NOMBRE_EXCEL_RESUMEN
    escribir_excel_resumen(destino, certificados=datos["certificado"])
    return destino


def procesar_carpeta_pdfs(carpeta: Path | str) -> dict[str, list[dict[str, Any]]]:
    carpeta = Path(carpeta)
    pdfs = listar_pdfs_carpeta_cuit(carpeta)
    return procesar_lista_pdfs(pdfs)


def _fila_certificado_a_lista(fila: dict[str, Any]) -> list[Any]:
    return [
        fila.get("Fecha emision"),
        fila.get("COE"),
        fila.get("Campaña"),
        fila.get("Tipo certificado"),
        fila.get("Grano y tipo"),
        fila.get("Depositario Razon Social"),
        fila.get("Depositario CUIT"),
        fila.get("Depositante Razon social"),
        fila.get("Depositante CUIT"),
        fila.get("Peso Bruto"),
        fila.get("Volatil"),
        fila.get("Secado"),
        fila.get("Zarandeo"),
        fila.get("Peso Neto"),
        fila.get("Zarandeo_serv"),
        fila.get("Secado_serv"),
        fila.get("Otros"),
        fila.get("Gastos Generales"),
        fila.get("Importe IVA"),
        fila.get("Cptos. No Gravados"),
        fila.get("Percepciones IVA"),
        fila.get("Otras Pecepciones"),
        fila.get("Total"),
        fila.get("CUIT Receptor"),
        fila.get("Razon Social Receptor"),
    ]


def _escribir_hoja_certificados(ws, filas: list[dict[str, Any]]) -> None:
    ws.append(_COLS_CERTIFICADOS)
    negrita = Font(bold=True)
    for cell in ws[1]:
        cell.font = negrita
    for fila in filas:
        ws.append(_fila_certificado_a_lista(fila))
    ws.freeze_panes = "A2"


def _escribir_hoja(ws, columnas: list[str], filas: list[dict[str, Any]]) -> None:
    ws.append(columnas)
    negrita = Font(bold=True)
    for cell in ws[1]:
        cell.font = negrita
    for fila in filas:
        ws.append([fila.get(c) for c in columnas])
    ws.freeze_panes = "A2"


def escribir_excel_resumen(
    destino: io.BytesIO | Path | str,
    *,
    certificados: list[dict[str, Any]] | None = None,
    lpg: list[dict[str, Any]] | None = None,
    hacienda: list[dict[str, Any]] | None = None,
) -> None:
    del lpg, hacienda  # LPG y hacienda ya no se incluyen en el Excel resumen VL.
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_CERTIFICADOS
    _escribir_hoja_certificados(ws, certificados or [])

    if isinstance(destino, io.BytesIO):
        destino.seek(0)
        destino.truncate(0)
        wb.save(destino)
        destino.seek(0)
    else:
        wb.save(destino)


def procesar_carpeta_a_excel(carpeta: Path | str, destino: Path | str | io.BytesIO) -> dict[str, int]:
    pdfs = listar_pdfs_carpeta_cuit(carpeta)
    datos = procesar_lista_pdfs(pdfs, solo_certificados=True)
    escribir_excel_resumen(destino, certificados=datos["certificado"])
    return {"certificados": len(datos["certificado"])}


if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(
        description="Extrae datos de PDF de liquidaciones/certificados a Excel resumen."
    )
    parser.add_argument(
        "carpeta",
        type=Path,
        help="Carpeta con archivos .pdf a procesar",
    )
    parser.add_argument(
        "-o",
        "--salida",
        type=Path,
        default=None,
        help="Excel de salida (por defecto: resumen_liquidaciones.xlsx en la carpeta)",
    )
    args = parser.parse_args()
    dest = args.salida or (args.carpeta / "resumen_liquidaciones.xlsx")
    meta = procesar_carpeta_a_excel(args.carpeta, dest)
    print(f"Excel generado: {dest}")
    print(f"Certificados: {meta['certificados']}")
