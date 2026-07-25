"""Cruce entre Libro IVA Compras (LIC) y Mis Comprobantes Recibidos (MCR)."""

from __future__ import annotations

import io
import re
import unicodedata
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils.dataframe import dataframe_to_rows

from sumar_imp_total import (
    _COLUMNA_COD_IMPUTACION_COMPROBANTES,
    _COLUMNA_NOM_IMPUTACION_COMPROBANTES,
    _aplicar_hoja_comprobantes_excel,
    _normalizar_clave_cuit_doc,
    agregar_columnas_imputacion_a_dataframe_comprobantes,
    parsear_numero_importe,
)

TOLERANCIA_TOTAL_CRUCE = 3.0
_SHEET_CRUCE = "No tomados en IVA"
_SHEET_LIC = "Libro IVA compras"
_SHEET_MCR = "Mis Comprobantes Recibidos"

_ROLES_CRUCE = ("punto_venta", "numero", "cuit", "total")
_ETIQUETAS_ROL = {
    "punto_venta": "punto de venta / PV",
    "numero": "número de comprobante / NC",
    "cuit": "CUIT del emisor",
    "total": "importe total del comprobante",
}


def _norm_encabezado(texto: object) -> str:
    s = unicodedata.normalize("NFKD", str(texto or ""))
    s = "".join(c for c in s if not unicodedata.combining(c))
    return re.sub(r"\s+", " ", s.strip().lower())


def _norm_entero_comprobante(val: object) -> str:
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return ""
    if isinstance(val, (int, float)) and not isinstance(val, bool):
        n = int(round(float(val)))
        return str(n)
    s = re.sub(r"\D", "", str(val).strip())
    if not s:
        return ""
    return str(int(s))


def _puntaje_columna(nombre: str, rol: str) -> float:
    h = _norm_encabezado(nombre)
    if not h:
        return -1.0
    if rol == "punto_venta":
        if re.search(r"punto\s*(de\s*)?venta|pto\.?\s*vta", h):
            return 10.0
        if h in ("pv", "p v", "p.v.", "p.v"):
            return 9.0
        if re.search(r"\bpv\b", h) and "iva" not in h:
            return 8.0
        if "punto" in h and "venta" in h:
            return 8.0
        return -1.0
    if rol == "numero":
        if re.search(
            r"numero\s*desde|nro\.?\s*desde|n[uú]mero\s*comprob|nro\.?\s*comprob",
            h,
        ):
            return 10.0
        if re.search(r"\bnc\b|num\.?\s*comp|nro\.?\s*comp|n[uú]m\.?\s*comp", h):
            return 9.0
        if re.search(r"numero|nro\.?\s*comp", h) and "hasta" not in h:
            return 7.0
        return -1.0
    if rol == "cuit":
        if re.search(r"tipo\s*doc|tipo\s*documento", h):
            return -1.0
        if re.search(r"nro\.?\s*doc\.?\s*emisor|numero\s*doc\.?\s*emisor", h):
            return 12.0
        if re.search(r"doc\.?\s*emisor|cuit.*emisor", h) and "receptor" not in h:
            return 10.0
        if h == "cuit" or h.startswith("cuit ") or h.endswith(" cuit"):
            return 9.0
        if re.search(r"cuit|c\.u\.i\.t", h):
            return 8.0
        if "documento" in h and "receptor" not in h:
            return 5.0
        return -1.0
    if rol == "total":
        if re.search(r"imp\.?\s*total|importe\s*total|total\s*comprob", h):
            return 10.0
        if re.search(r"^total$|total\s*fact|monto\s*total", h):
            return 8.0
        if re.search(r"importe|monto", h) and "iva" not in h and "neto" not in h:
            return 6.0
        if h == "total" or h.endswith(" total"):
            return 5.0
        if "total iva" in h or h.startswith("iva "):
            return -1.0
        if "total" in h and "neto" not in h:
            return 3.0
        return -1.0
    return -1.0


def detectar_columnas_cruce(df: pd.DataFrame) -> dict[str, str]:
    """
    Resuelve columnas mínimas para cruce: PV, número, CUIT emisor e importe total.
    No exige formato ARCA completo; basta con encabezados reconocibles en cada archivo.
    """
    mejor: dict[str, tuple[float, str]] = {r: (-1.0, "") for r in _ROLES_CRUCE}
    for col in df.columns:
        nombre = str(col)
        for rol in _ROLES_CRUCE:
            score = _puntaje_columna(nombre, rol)
            if score > mejor[rol][0]:
                mejor[rol] = (score, nombre)
    faltantes = [r for r in _ROLES_CRUCE if mejor[r][0] < 0]
    if faltantes:
        nombres = ", ".join(str(c) for c in df.columns)
        falt_txt = ", ".join(_ETIQUETAS_ROL[r] for r in faltantes)
        raise ValueError(
            f"No se detectaron columnas para el cruce ({falt_txt}). "
            f"Columnas del archivo: {nombres}"
        )
    elegidas = {r: mejor[r][1] for r in _ROLES_CRUCE}
    if len(set(elegidas.values())) < len(_ROLES_CRUCE):
        raise ValueError(
            "Las columnas detectadas para el cruce se solapan; "
            "revisá que PV, número, CUIT y total tengan encabezados distintos."
        )
    return elegidas


def _leer_excel_para_cruce(entrada, *, ui_lang: str = "es") -> pd.DataFrame:
    """
    Lee .xlsx probando varias filas de encabezado.
    Elige la primera donde se detecten PV, número, CUIT y total.
    """
    ultimo_error = ""
    for header_row in range(0, 5):
        if hasattr(entrada, "seek"):
            entrada.seek(0)
        try:
            raw = pd.read_excel(entrada, header=header_row)
        except Exception:
            continue
        raw.columns = raw.columns.astype(str).str.strip()
        if raw.empty or len(raw.columns) < 4:
            continue
        try:
            detectar_columnas_cruce(raw)
            return raw.copy()
        except ValueError as exc:
            ultimo_error = str(exc)
    msg = ultimo_error or (
        "No se pudo leer el archivo Excel."
        if ui_lang == "es"
        else "Could not read the Excel file."
    )
    raise ValueError(msg)


def leer_archivo_para_cruce(
    entrada,
    *,
    nombre_archivo: str | None = None,
    ui_lang: str = "es",
) -> pd.DataFrame:
    """LIC o MCR: solo se requieren columnas identificables de PV, NC, CUIT y total."""
    nombre = (nombre_archivo or "").lower()
    if nombre.endswith(".csv"):
        return _leer_csv_para_cruce(entrada, ui_lang=ui_lang)
    return _leer_excel_para_cruce(entrada, ui_lang=ui_lang)


def _leer_csv_para_cruce(entrada, *, ui_lang: str = "es") -> pd.DataFrame:
    ultimo_error = ""
    for kwargs in (
        {"sep": ";", "encoding": "utf-8-sig"},
        {"sep": ",", "encoding": "utf-8-sig"},
        {"sep": ";", "encoding": "latin-1"},
        {"sep": ",", "encoding": "latin-1"},
    ):
        for header_row in (0, 1):
            if hasattr(entrada, "seek"):
                entrada.seek(0)
            try:
                raw = pd.read_csv(entrada, header=header_row, **kwargs)
            except Exception:
                continue
            raw.columns = raw.columns.astype(str).str.strip()
            if raw.empty or len(raw.columns) < 4:
                continue
            try:
                detectar_columnas_cruce(raw)
                return raw.copy()
            except ValueError as exc:
                ultimo_error = str(exc)
    msg = ultimo_error or (
        "No se pudo leer el CSV."
        if ui_lang == "es"
        else "Could not read the CSV file."
    )
    raise ValueError(msg)


def _normalizar_cuit_cruce(val: object) -> str:
    """
    CUIT/DNI del emisor en formato comparable: solo dígitos, sin - . ni espacios.
    Acepta 11 dígitos corridos, 30-68898047-6, 30.68898047.6, etc.
    """
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return ""
    if isinstance(val, str):
        s = val.strip()
        if not s:
            return ""
        if re.fullmatch(r"\d+\.0+", s):
            try:
                val = int(float(s))
            except ValueError:
                pass
    digits = _normalizar_clave_cuit_doc(val)
    if not digits:
        return ""
    if len(digits) > 11 and digits.startswith("0"):
        recortado = digits.lstrip("0")
        if len(recortado) == 11:
            return recortado
    return digits


def _identidad_comprobante(
    fila: pd.Series, cols: dict[str, str]
) -> tuple[str, str, str, float] | None:
    """PV + número + CUIT + importe total (normalizados)."""
    pv = _norm_entero_comprobante(fila.get(cols["punto_venta"]))
    num = _norm_entero_comprobante(fila.get(cols["numero"]))
    cuit = _normalizar_cuit_cruce(fila.get(cols["cuit"]))
    total = float(parsear_numero_importe(fila.get(cols["total"])) or 0.0)
    if not pv or not num or not cuit:
        return None
    return (pv, num, cuit, total)


def _lista_comprobantes_lic(
    df: pd.DataFrame, cols: dict[str, str]
) -> list[tuple[str, str, str, float]]:
    out: list[tuple[str, str, str, float]] = []
    for _, fila in df.iterrows():
        ident = _identidad_comprobante(fila, cols)
        if ident is not None:
            out.append(ident)
    return out


def _comprobante_coincide_en_lic(
    ident_mcr: tuple[str, str, str, float],
    comprobantes_lic: list[tuple[str, str, str, float]],
    *,
    tolerancia: float = TOLERANCIA_TOTAL_CRUCE,
) -> bool:
    """
    Coincide si en LIC hay un comprobante con el mismo PV, número y CUIT emisor
    y un importe total que difiere en hasta ``tolerancia`` pesos.
    Si los tres primeros coinciden pero el total difiere más, NO hay match.
    """
    pv, num, cuit, total_mcr = ident_mcr
    for pv_l, num_l, cuit_l, total_lic in comprobantes_lic:
        if pv == pv_l and num == num_l and cuit == cuit_l:
            if abs(total_mcr - total_lic) <= tolerancia:
                return True
    return False


def filtrar_mcr_sin_cruzar(
    df_mcr: pd.DataFrame,
    df_lic: pd.DataFrame,
    *,
    tolerancia: float = TOLERANCIA_TOTAL_CRUCE,
) -> pd.DataFrame:
    """Filas de MCR sin par en LIC (PV + NC + CUIT + total con tolerancia en el importe)."""
    cols_mcr = detectar_columnas_cruce(df_mcr)
    cols_lic = detectar_columnas_cruce(df_lic)
    comprobantes_lic = _lista_comprobantes_lic(df_lic, cols_lic)

    indices: list[int] = []
    for i, fila in df_mcr.iterrows():
        ident = _identidad_comprobante(fila, cols_mcr)
        if ident is None:
            indices.append(int(i))
            continue
        if not _comprobante_coincide_en_lic(
            ident, comprobantes_lic, tolerancia=tolerancia
        ):
            indices.append(int(i))

    if not indices:
        return df_mcr.iloc[0:0].copy()
    return df_mcr.loc[indices].copy()


def _mapa_imputacion_claves_normalizadas(
    mapa: dict[str, tuple[str, str]],
) -> dict[str, tuple[str, str]]:
    """Unifica claves del mapa de imputación (CUIT con/sin guiones, ceros, etc.)."""
    out: dict[str, tuple[str, str]] = {}
    for clave, valor in mapa.items():
        k = _normalizar_cuit_cruce(clave) or _normalizar_clave_cuit_doc(clave)
        if k and k not in out:
            out[k] = valor
    return out


def agregar_imputacion_a_dataframe_cruce(
    df: pd.DataFrame,
    mapa_imputaciones: dict[str, tuple[str, str]] | None,
) -> pd.DataFrame:
    """
    Añade ``Cód. imputación`` e ``Imputación contable`` según el CUIT emisor de cada fila.
    Si el archivo tiene columna ARCA ``Nro. Doc. Emisor``, reutiliza la lógica del procesador;
    si no, usa la columna CUIT detectada para el cruce.
    """
    if not mapa_imputaciones:
        return df
    mapa = _mapa_imputacion_claves_normalizadas(mapa_imputaciones)
    if "Nro. Doc. Emisor" in df.columns:
        return agregar_columnas_imputacion_a_dataframe_comprobantes(
            df, mapa, emitidos=False
        )
    cols = detectar_columnas_cruce(df)
    col_cuit = cols["cuit"]
    out = df.copy()
    codigos: list[str] = []
    nombres: list[str] = []
    for val in out[col_cuit]:
        k = _normalizar_cuit_cruce(val)
        if not k or k not in mapa:
            codigos.append("")
            nombres.append("")
        else:
            c, n = mapa[k]
            codigos.append(c)
            nombres.append(n)
    out[_COLUMNA_COD_IMPUTACION_COMPROBANTES] = codigos
    out[_COLUMNA_NOM_IMPUTACION_COMPROBANTES] = nombres
    return out


def _escribir_hoja_datos(
    wb: Workbook, df: pd.DataFrame, titulo: str, *, indice: int | None = None
) -> None:
    if indice is None:
        ws = wb.create_sheet(titulo)
    else:
        ws = wb.create_sheet(titulo, indice)
    for row in dataframe_to_rows(df, index=False, header=True):
        ws.append(row)
    negrita = Font(bold=True)
    for cell in ws[1]:
        cell.font = negrita
    ws.freeze_panes = "A2"
    if ws.max_row >= 1 and ws.max_column >= 1:
        from openpyxl.utils import get_column_letter

        ult = get_column_letter(ws.max_column)
        ws.auto_filter.ref = f"A1:{ult}{ws.max_row}"


def _es_export_arca_comprobantes(df: pd.DataFrame) -> bool:
    cols = {_norm_encabezado(c) for c in df.columns}
    return "imp. total" in cols or "imp total" in cols


def escribir_excel_cruce_lic_mcr(
    destino: io.BytesIO | Path | str,
    *,
    df_cruce: pd.DataFrame,
    df_lic: pd.DataFrame,
    df_mcr: pd.DataFrame,
) -> None:
    """Genera libro con Cruce (1.ª hoja), LIC original y MCR original."""
    wb = Workbook()
    ws_cruce = wb.active
    ws_cruce.title = _SHEET_CRUCE
    for row in dataframe_to_rows(df_cruce, index=False, header=True):
        ws_cruce.append(row)
    encab = [c.value for c in ws_cruce[1]]
    if _es_export_arca_comprobantes(df_cruce):
        _aplicar_hoja_comprobantes_excel(
            wb, ws_cruce, encab, emitidos=False, titulo_hoja=_SHEET_CRUCE
        )
    else:
        negrita = Font(bold=True)
        for cell in ws_cruce[1]:
            cell.font = negrita
        ws_cruce.freeze_panes = "A2"
        if ws_cruce.max_row >= 1 and ws_cruce.max_column >= 1:
            from openpyxl.utils import get_column_letter

            ult = get_column_letter(ws_cruce.max_column)
            ws_cruce.auto_filter.ref = f"A1:{ult}{ws_cruce.max_row}"

    _escribir_hoja_datos(wb, df_lic, _SHEET_LIC)
    _escribir_hoja_datos(wb, df_mcr, _SHEET_MCR)

    if isinstance(destino, io.BytesIO):
        destino.seek(0)
        destino.truncate(0)
        wb.save(destino)
        destino.seek(0)
    else:
        wb.save(destino)


def procesar_cruce_lic_mcr(
    lic_bytes: bytes,
    mcr_bytes: bytes,
    *,
    lic_nombre: str,
    mcr_nombre: str,
    ui_lang: str = "es",
    mapa_imputaciones: dict[str, tuple[str, str]] | None = None,
) -> tuple[bytes, dict[str, Any]]:
    buf_lic = io.BytesIO(lic_bytes)
    buf_mcr = io.BytesIO(mcr_bytes)

    df_lic = leer_archivo_para_cruce(buf_lic, nombre_archivo=lic_nombre, ui_lang=ui_lang)
    df_mcr = leer_archivo_para_cruce(buf_mcr, nombre_archivo=mcr_nombre, ui_lang=ui_lang)

    df_cruce = filtrar_mcr_sin_cruzar(df_mcr, df_lic)
    if mapa_imputaciones:
        df_cruce = agregar_imputacion_a_dataframe_cruce(df_cruce, mapa_imputaciones)
    salida = io.BytesIO()
    escribir_excel_cruce_lic_mcr(
        salida,
        df_cruce=df_cruce,
        df_lic=df_lic,
        df_mcr=df_mcr,
    )
    meta = {
        "total_mcr": int(len(df_mcr)),
        "total_lic": int(len(df_lic)),
        "total_cruce": int(len(df_cruce)),
        "con_imputacion": bool(mapa_imputaciones),
    }
    return salida.getvalue(), meta
