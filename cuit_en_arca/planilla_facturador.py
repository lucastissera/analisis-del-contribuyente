"""Lectura de planilla Excel Facturador (Comprobantes en línea)."""

from __future__ import annotations

import io
import re
import unicodedata
from dataclasses import dataclass
from datetime import date

from openpyxl import load_workbook

from cuit_en_arca.errores import CredencialesArchivoError
from cuit_en_arca.planilla_lote import _celda_str, _norm_header, _solo_digitos
from cuit_en_arca.plantillas_importacion import ruta_plantilla_facturador_excel
from cuit_en_arca.validacion import parsear_fecha_argentina


@dataclass(frozen=True)
class FilaPlanillaFacturador:
    fila_excel: int
    cuit_login: str
    clave_fiscal: str
    representado: str
    punto_venta: str
    tipo_comprobante: str
    fecha_comprobante: str
    concepto: str
    moneda_extranjera: str
    per_facturado_desde: str
    per_facturado_hasta: str
    vto_pago: str
    condicion_iva: str
    tipo_documento: str
    nro_documento: str
    condiciones_venta: str
    producto_servicio: str
    cantidad: str
    unidad_medida: str
    precio_unitario: str


_COLUMNAS: dict[str, tuple[str, ...]] = {
    "cuit": ("cuit",),
    "clave": ("clave fiscal",),
    "representado": ("representado",),
    "punto_venta": ("punto venta",),
    "tipo_comprobante": ("tipo comprobante",),
    "fecha_comprobante": ("fecha comprobante",),
    "concepto": ("concepto",),
    "moneda_extranjera": ("moneda extranjera",),
    "per_desde": ("per facturado desde", "periodo facturado desde"),
    "per_hasta": ("per facturado hasta", "periodo facturado hasta"),
    "vto_pago": ("vto. para el pago", "vto para el pago", "vencimiento"),
    "condicion_iva": ("condicion frente al iva", "condición frente al iva"),
    "tipo_documento": ("tipo documento",),
    "nro_documento": ("n° documento", "nº documento", "nro documento", "numero documento"),
    "condiciones_venta": ("condiciones de venta",),
    "producto_servicio": ("producto/servicio", "producto servicio"),
    "cantidad": ("cantidad",),
    "unidad_medida": ("unidad medida",),
    "precio_unitario": ("precio unitario",),
}


def _detectar_columnas(headers: list[str]) -> dict[str, int | None]:
    idx: dict[str, int | None] = {k: None for k in _COLUMNAS}
    for i, h in enumerate(headers):
        if not h:
            continue
        for clave, variantes in _COLUMNAS.items():
            if idx[clave] is not None:
                continue
            if any(v in h for v in variantes):
                idx[clave] = i
    if idx["cuit"] is None:
        idx["cuit"] = 0
    if idx["clave"] is None:
        idx["clave"] = 1
    if idx["representado"] is None:
        idx["representado"] = 2
    return idx


def _parece_instructivo(row_vals: tuple, cols: dict[str, int | None]) -> bool:
    if any(
        _celda_str(row_vals[i]) if i is not None and i < len(row_vals) else ""
        for i in (cols.get("cuit"), cols.get("clave"), cols.get("representado"))
        if i is not None
    ):
        return False
    texto = " ".join(_celda_str(v).lower() for v in row_vals if v is not None)
    return "instructivo" in texto or ("11 d" in texto and "cuit" in texto)


def _get(row_vals: tuple, cols: dict[str, int | None], key: str) -> str:
    i = cols.get(key)
    if i is None or i >= len(row_vals):
        return ""
    return _celda_str(row_vals[i])


def _parsear_fila(fila_num: int, row_vals: tuple, cols: dict[str, int | None]) -> FilaPlanillaFacturador | None:
    if _parece_instructivo(row_vals, cols):
        return None

    cuit = _get(row_vals, cols, "cuit")
    clave_fiscal = _get(row_vals, cols, "clave")
    representado = _get(row_vals, cols, "representado")
    if not any((cuit, clave_fiscal, representado, _get(row_vals, cols, "tipo_comprobante"))):
        return None

    if not cuit or not clave_fiscal:
        raise CredencialesArchivoError(
            f"Fila {fila_num}: faltan CUIT de ingreso o clave fiscal."
        )
    if not representado:
        raise CredencialesArchivoError(
            f"Fila {fila_num}: falta el representado (razón social, no CUIT)."
        )

    cuit_norm = _solo_digitos(cuit, f"Fila {fila_num} CUIT")
    condicion_iva = _get(row_vals, cols, "condicion_iva")
    nro_doc = _get(row_vals, cols, "nro_documento")
    campos_obligatorios = (
        ("punto_venta", "Punto Venta"),
        ("tipo_comprobante", "Tipo Comprobante"),
        ("concepto", "Concepto"),
        ("condicion_iva", "Condición frente al IVA"),
        ("tipo_documento", "Tipo Documento"),
        ("condiciones_venta", "Condiciones de venta"),
        ("producto_servicio", "Producto/Servicio"),
        ("cantidad", "Cantidad"),
        ("precio_unitario", "Precio Unitario"),
    )
    for campo_key, etiqueta in campos_obligatorios:
        if not _get(row_vals, cols, campo_key):
            raise CredencialesArchivoError(f"Fila {fila_num}: falta {etiqueta}.")
    if not nro_doc and not es_consumidor_final(condicion_iva):
        raise CredencialesArchivoError(
            f"Fila {fila_num}: falta N° Documento (puede quedar vacío solo con Consumidor Final)."
        )

    fecha_comp = _get(row_vals, cols, "fecha_comprobante")
    if fecha_comp and not parsear_fecha_argentina(fecha_comp):
        raise CredencialesArchivoError(
            f"Fila {fila_num}: Fecha Comprobante inválida (use dd/mm/aaaa)."
        )

    return FilaPlanillaFacturador(
        fila_excel=fila_num,
        cuit_login=cuit_norm,
        clave_fiscal=clave_fiscal,
        representado=representado.strip(),
        punto_venta=_get(row_vals, cols, "punto_venta"),
        tipo_comprobante=_get(row_vals, cols, "tipo_comprobante"),
        fecha_comprobante=fecha_comp,
        concepto=_get(row_vals, cols, "concepto"),
        moneda_extranjera=_get(row_vals, cols, "moneda_extranjera"),
        per_facturado_desde=_get(row_vals, cols, "per_desde"),
        per_facturado_hasta=_get(row_vals, cols, "per_hasta"),
        vto_pago=_get(row_vals, cols, "vto_pago"),
        condicion_iva=condicion_iva,
        tipo_documento=_get(row_vals, cols, "tipo_documento"),
        nro_documento=nro_doc,
        condiciones_venta=_get(row_vals, cols, "condiciones_venta"),
        producto_servicio=_get(row_vals, cols, "producto_servicio"),
        cantidad=_get(row_vals, cols, "cantidad"),
        unidad_medida=normalizar_unidad_medida(_get(row_vals, cols, "unidad_medida")),
        precio_unitario=_get(row_vals, cols, "precio_unitario"),
    )


def leer_planilla_facturador_con_errores(
    fuente: str | bytes | io.BytesIO,
) -> tuple[list[FilaPlanillaFacturador], list[str]]:
    if isinstance(fuente, (bytes, bytearray)):
        stream: io.BytesIO | str = io.BytesIO(fuente)
    else:
        stream = fuente

    wb = load_workbook(stream, read_only=True, data_only=True)
    try:
        ws = wb["Formato"] if "Formato" in wb.sheetnames else wb.active
        filas: list[FilaPlanillaFacturador] = []
        errores: list[str] = []
        cols: dict[str, int | None] | None = None

        for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
            row_vals = tuple(row)
            if i == 1:
                headers = [_norm_header(v) for v in row_vals]
                cols = _detectar_columnas(headers)
                continue
            if cols is None:
                break
            try:
                fila = _parsear_fila(i, row_vals, cols)
            except CredencialesArchivoError as exc:
                errores.append(str(exc))
                continue
            if fila is not None:
                filas.append(fila)
        return filas, errores
    finally:
        wb.close()


def es_consumidor_final(condicion_iva: str) -> bool:
    return "consumidor final" in normalizar_texto_arca(condicion_iva)


def normalizar_unidad_medida(valor: str) -> str:
    """Vacío o «SIN DESCRIPCION» son equivalentes para ARCA."""
    if not (valor or "").strip():
        return "SIN DESCRIPCION"
    if normalizar_texto_arca(valor) in ("sin descripcion",):
        return "SIN DESCRIPCION"
    return valor.strip()


def normalizar_texto_arca(texto: str) -> str:
    s = (texto or "").strip().lower()
    s = unicodedata.normalize("NFD", s)
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    s = re.sub(r"\s+", " ", s)
    return s.strip()


EXPRESS_VALORES_FIJOS = {
    "moneda_extranjera": "NO",
    "condicion_iva": "Consumidor Final",
    "tipo_documento": "CUIT",
    "nro_documento": "",
    "condiciones_venta": "Contado",
    "unidad_medida": "SIN DESCRIPCION",
}

_CONCEPTOS_EXPRESS = frozenset({"servicios", "productos"})
_RE_CANTIDAD_EXPRESS = re.compile(r"^\d+(?:[.,]\d{1,5})?$")
_RE_PRECIO_EXPRESS = re.compile(r"^\d+(?:[.,]\d+)?$")


def _fecha_hoy_argentina() -> str:
    return date.today().strftime("%d/%m/%Y")


def listar_tipos_comprobante_modelo() -> list[str]:
    """Tipos de comprobante de la hoja Tablas del Excel modelo."""
    ruta = ruta_plantilla_facturador_excel()
    if not ruta.is_file():
        return []
    wb = load_workbook(ruta, read_only=True, data_only=True)
    try:
        if "Tablas" not in wb.sheetnames:
            return []
        ws = wb["Tablas"]
        filas = list(ws.iter_rows(values_only=True))
        if not filas:
            return []
        headers = [normalizar_texto_arca(str(c or "")) for c in filas[0]]
        col = None
        for i, h in enumerate(headers):
            if "tipos comprobante" in h or h == "tipo comprobante":
                col = i
                break
        if col is None:
            return []
        vistos: set[str] = set()
        resultado: list[str] = []
        for row in filas[1:]:
            if col >= len(row):
                continue
            valor = _celda_str(row[col])
            if not valor:
                continue
            clave = normalizar_texto_arca(valor)
            if clave in vistos:
                continue
            vistos.add(clave)
            resultado.append(valor)
        return resultado
    finally:
        wb.close()


def _normalizar_numero_decimal(valor: str, *, etiqueta: str, fila_num: int) -> str:
    texto = (valor or "").strip().replace(" ", "")
    if not texto:
        raise CredencialesArchivoError(f"Fila express {fila_num}: falta {etiqueta}.")
    if not _RE_CANTIDAD_EXPRESS.match(texto) if etiqueta == "Cantidad" else not _RE_PRECIO_EXPRESS.match(texto):
        raise CredencialesArchivoError(
            f"Fila express {fila_num}: {etiqueta} inválido{f' (máx. 5 decimales)' if etiqueta == 'Cantidad' else ''}."
        )
    return texto.replace(",", ".")


def construir_filas_express(
    *,
    cuit_login: str,
    clave_fiscal: str,
    representado: str,
    punto_venta: str,
    fechas: list[str],
    tipos: list[str],
    conceptos: list[str],
    productos: list[str],
    cantidades: list[str],
    precios: list[str],
) -> tuple[list[FilaPlanillaFacturador], list[str]]:
    """Arma filas de emisión express (misma sesión CUIT / representado / punto de venta)."""
    errores: list[str] = []
    cuit_norm = ""
    try:
        cuit_norm = _solo_digitos(cuit_login, "CUIT ingreso")
    except CredencialesArchivoError as exc:
        errores.append(str(exc))
    clave = (clave_fiscal or "").strip()
    pv = (punto_venta or "").strip()
    if not clave:
        errores.append("Falta la clave fiscal.")
    if not pv:
        errores.append("Falta el Punto de Venta.")

    tipos_validos = {normalizar_texto_arca(t) for t in listar_tipos_comprobante_modelo()}
    n = max(len(fechas), len(tipos), len(conceptos), len(productos), len(cantidades), len(precios))
    filas: list[FilaPlanillaFacturador] = []

    for i in range(n):
        fila_num = i + 1
        fecha = (fechas[i] if i < len(fechas) else "").strip()
        tipo = (tipos[i] if i < len(tipos) else "").strip()
        concepto = (conceptos[i] if i < len(conceptos) else "").strip()
        producto = (productos[i] if i < len(productos) else "").strip()
        cantidad = (cantidades[i] if i < len(cantidades) else "").strip()
        precio = (precios[i] if i < len(precios) else "").strip()

        if not any((tipo, concepto, producto, cantidad, precio)):
            continue
        if not cuit_norm or not clave or not pv:
            continue

        try:
            if not tipo:
                raise CredencialesArchivoError(f"Fila express {fila_num}: falta Tipo de comprobante.")
            if tipos_validos and normalizar_texto_arca(tipo) not in tipos_validos:
                raise CredencialesArchivoError(
                    f"Fila express {fila_num}: tipo de comprobante «{tipo}» no figura en el modelo."
                )
            if not concepto:
                raise CredencialesArchivoError(f"Fila express {fila_num}: falta Concepto.")
            concepto_norm = normalizar_texto_arca(concepto)
            if concepto_norm not in _CONCEPTOS_EXPRESS:
                raise CredencialesArchivoError(
                    f"Fila express {fila_num}: Concepto debe ser Servicios o Productos."
                )
            if not producto:
                raise CredencialesArchivoError(f"Fila express {fila_num}: falta Producto/Servicio.")
            cantidad_norm = _normalizar_numero_decimal(cantidad, etiqueta="Cantidad", fila_num=fila_num)
            precio_norm = _normalizar_numero_decimal(precio, etiqueta="Precio unitario", fila_num=fila_num)

            fecha_comp = fecha or _fecha_hoy_argentina()
            if fecha and not parsear_fecha_argentina(fecha):
                raise CredencialesArchivoError(
                    f"Fila express {fila_num}: Fecha inválida (use dd/mm/aaaa)."
                )

            per_desde = fecha_comp if concepto_norm == "servicios" else ""
            per_hasta = fecha_comp if concepto_norm == "servicios" else ""
            vto = fecha_comp if concepto_norm == "servicios" else ""

            filas.append(
                FilaPlanillaFacturador(
                    fila_excel=fila_num,
                    cuit_login=cuit_norm,
                    clave_fiscal=clave,
                    representado=(representado or "").strip(),
                    punto_venta=pv,
                    tipo_comprobante=tipo,
                    fecha_comprobante=fecha_comp,
                    concepto=concepto,
                    moneda_extranjera=EXPRESS_VALORES_FIJOS["moneda_extranjera"],
                    per_facturado_desde=per_desde,
                    per_facturado_hasta=per_hasta,
                    vto_pago=vto,
                    condicion_iva=EXPRESS_VALORES_FIJOS["condicion_iva"],
                    tipo_documento=EXPRESS_VALORES_FIJOS["tipo_documento"],
                    nro_documento=EXPRESS_VALORES_FIJOS["nro_documento"],
                    condiciones_venta=EXPRESS_VALORES_FIJOS["condiciones_venta"],
                    producto_servicio=producto,
                    cantidad=cantidad_norm,
                    unidad_medida=EXPRESS_VALORES_FIJOS["unidad_medida"],
                    precio_unitario=precio_norm,
                )
            )
        except CredencialesArchivoError as exc:
            errores.append(str(exc))

    if not filas and not errores:
        errores.append("Completá al menos una factura express con todos sus campos.")
    return filas, errores
