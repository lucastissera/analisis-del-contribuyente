"""Lectura de planilla Excel Ventas y Liquidaciones (VyL)."""

from __future__ import annotations

import io
from dataclasses import dataclass

from openpyxl import load_workbook

from cuit_en_arca.errores import CredencialesArchivoError
from cuit_en_arca.planilla_lote import (
    _celda_rango,
    _celda_str,
    _detectar_columnas,
    _norm_header,
    _solo_digitos,
)
from cuit_en_arca.validacion import parsear_fecha_argentina, validar_rango_max_un_anio


@dataclass(frozen=True)
class FilaPlanillaVl:
    fila_excel: int
    cuit_login: str
    clave_fiscal: str
    nombre_representado: str
    fecha_desde: str
    fecha_hasta: str


def _parece_fila_instructivo(row_vals: tuple, cols: dict) -> bool:
    """Filas de ayuda del Excel (columnas G/H). No saltear filas con datos en A–E."""

    def _celda_col(key: str) -> str:
        i = cols.get(key)
        if i is None or i >= len(row_vals):
            return ""
        return _celda_str(row_vals[i])

    if any(
        _celda_col(k)
        for k in ("login", "clave", "repr", "desde", "hasta", "rango")
    ):
        return False

    texto = " ".join(_celda_str(v).lower() for v in row_vals if v is not None)
    return "instructivo" in texto or ("11 d" in texto and "cuit" in texto)


def _parsear_fila_vl(fila_num: int, row_vals: tuple, cols: dict) -> FilaPlanillaVl | None:
    def get(key: str) -> str:
        i = cols[key]
        if i is None or i >= len(row_vals):
            return ""
        return _celda_str(row_vals[i])

    cuit_log = get("login")
    clave = get("clave")
    nombre = get("repr")
    i_rango = cols["rango"]
    raw_rango = (
        _celda_rango(row_vals[i_rango])
        if i_rango is not None and i_rango < len(row_vals)
        else ""
    )

    def get_fecha(key: str) -> str:
        i = cols.get(key)
        if i is None or i >= len(row_vals):
            return ""
        return _celda_rango(row_vals[i])

    if not cuit_log and not clave and not nombre and not raw_rango:
        return None
    if not cuit_log or not clave:
        raise CredencialesArchivoError(
            f"Fila {fila_num}: faltan CUIT de ingreso o clave fiscal."
        )
    if not nombre.strip():
        raise CredencialesArchivoError(
            f"Fila {fila_num}: falta el nombre del representado (columna Representado)."
        )

    par: tuple[str, str] | None = None
    if not raw_rango:
        fd_alt = get_fecha("desde")
        fh_alt = get_fecha("hasta")
        if fd_alt and fh_alt:
            par = (fd_alt, fh_alt)
        else:
            raise CredencialesArchivoError(
                f"Fila {fila_num}: faltan Fecha Liq Desde / Fecha Liq Hasta."
            )
    else:
        from cuit_en_arca.validacion import parsear_rango_fechas_texto

        par = parsear_rango_fechas_texto(raw_rango)
        if not par:
            fd_alt = get_fecha("desde")
            fh_alt = get_fecha("hasta")
            if fd_alt and fh_alt:
                par = (fd_alt, fh_alt)
    if not par:
        raise CredencialesArchivoError(
            f"Fila {fila_num}: rango de fechas inválido (use dd/mm/yyyy)."
        )
    fd, fh = par
    desde = parsear_fecha_argentina(fd)
    hasta = parsear_fecha_argentina(fh)
    validar_rango_max_un_anio(desde, hasta)
    cuit_login = _solo_digitos(cuit_log, f"CUIT ingreso fila {fila_num}")

    return FilaPlanillaVl(
        fila_excel=fila_num,
        cuit_login=cuit_login,
        clave_fiscal=clave,
        nombre_representado=nombre.strip(),
        fecha_desde=fd,
        fecha_hasta=fh,
    )


def leer_planilla_vl_con_errores(buf: io.BytesIO) -> tuple[list[FilaPlanillaVl], list[str]]:
    errores: list[str] = []
    try:
        buf.seek(0)
        wb = load_workbook(buf, read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
    except Exception as exc:
        raise CredencialesArchivoError(f"No se pudo leer la planilla VyL: {exc}") from exc

    if not rows:
        raise CredencialesArchivoError("La planilla VyL está vacía.")

    headers = [_norm_header(c) for c in rows[0]]
    cols = _detectar_columnas(headers)

    filas: list[FilaPlanillaVl] = []
    for i, row in enumerate(rows[1:], start=2):
        if not row or all(v is None or str(v).strip() == "" for v in row):
            continue
        if _parece_fila_instructivo(row, cols):
            continue
        try:
            parsed = _parsear_fila_vl(i, row, cols)
        except CredencialesArchivoError as exc:
            errores.append(str(exc))
            continue
        if parsed is not None:
            filas.append(parsed)

    if not filas and not errores:
        raise CredencialesArchivoError(
            "No hay filas de datos en la planilla VyL (revisá fila 2 en adelante)."
        )
    return filas, errores


def parsear_entradas_manuales_vl(
    cuits_login: list[str],
    claves: list[str],
    nombres_repr: list[str],
    desdes: list[str],
    hastas: list[str],
) -> tuple[list[FilaPlanillaVl], list[str]]:
    errores: list[str] = []
    filas: list[FilaPlanillaVl] = []

    def _at(lista, i):
        return (lista[i] if i < len(lista) else "").strip()

    n = max(len(cuits_login), len(claves), len(nombres_repr), len(desdes), len(hastas))
    cols = {
        "login": 0,
        "clave": 1,
        "repr": 2,
        "rango": None,
        "desde": 3,
        "hasta": 4,
    }

    for i in range(n):
        cuit_log = _at(cuits_login, i)
        clave = _at(claves, i)
        nombre = _at(nombres_repr, i)
        fd = _at(desdes, i)
        fh = _at(hastas, i)
        if not any((cuit_log, clave, nombre, fd, fh)):
            continue
        if not cuit_log or not clave or not nombre or not fd or not fh:
            errores.append(
                "Entrada manual: completá CUIT ingreso, clave, nombre representado y fechas."
            )
            continue
        try:
            cuit_login = _solo_digitos(cuit_log, "CUIT ingreso (manual)")
            desde = parsear_fecha_argentina(fd)
            hasta = parsear_fecha_argentina(fh)
            validar_rango_max_un_anio(desde, hasta)
        except Exception as exc:
            errores.append(f"Entrada manual fila {i + 1}: {exc}")
            continue
        filas.append(
            FilaPlanillaVl(
                fila_excel=i + 1,
                cuit_login=cuit_login,
                clave_fiscal=clave,
                nombre_representado=nombre,
                fecha_desde=fd,
                fecha_hasta=fh,
            )
        )
    return filas, errores
