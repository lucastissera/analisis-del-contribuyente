"""Lectura de planilla Excel con múltiples filas para descarga masiva ARCA."""

from __future__ import annotations

import io
import re
from dataclasses import dataclass
from datetime import date, datetime

from openpyxl import load_workbook

from cuit_en_arca.errores import CredencialesArchivoError
from cuit_en_arca.validacion import parsear_fecha_argentina, parsear_rango_fechas_texto, validar_rango_max_un_anio


@dataclass(frozen=True)
class FilaPlanillaArca:
    fila_excel: int
    cuit_login: str
    clave_fiscal: str
    cuit_representado: str
    fecha_desde: str
    fecha_hasta: str


def _solo_digitos(s: str, etiqueta: str) -> str:
    d = re.sub(r"\D", "", str(s))
    if len(d) != 11:
        raise CredencialesArchivoError(
            f"{etiqueta} (fila indicada): se esperaban 11 dígitos."
        )
    return d


def _norm_header(val) -> str:
    if val is None:
        return ""
    return re.sub(r"\s+", " ", str(val).strip().lower())


def _detectar_columnas(headers: list[str]) -> dict[str, int | None]:
    idx_login = idx_clave = idx_repr = idx_rango = idx_desde = idx_hasta = None
    for i, h in enumerate(headers):
        if not h:
            continue
        if idx_login is None and any(
            x in h for x in ("cuit ingreso", "cuit login", "cuit representante", "cuit ingres")
        ):
            idx_login = i
        elif idx_clave is None and "clave" in h and "fiscal" in h:
            idx_clave = i
        elif idx_repr is None and "representado" in h:
            idx_repr = i
        elif idx_rango is None and ("rango" in h and "fecha" in h):
            idx_rango = i
        elif idx_desde is None and "fecha" in h and any(x in h for x in ("desde", "inicio")):
            idx_desde = i
        elif idx_hasta is None and "fecha" in h and any(x in h for x in ("hasta", "fin")):
            idx_hasta = i
    if idx_login is None and idx_clave is None and idx_repr is None and idx_rango is None:
        return {
            "login": 0,
            "clave": 1,
            "repr": 2,
            "rango": 3,
            "desde": 4 if len(headers) > 4 else None,
            "hasta": 5 if len(headers) > 5 else None,
        }
    return {
        "login": idx_login if idx_login is not None else 0,
        "clave": idx_clave if idx_clave is not None else 1,
        "repr": idx_repr if idx_repr is not None else 2,
        "rango": idx_rango if idx_rango is not None else 3,
        "desde": idx_desde,
        "hasta": idx_hasta,
    }


def _celda_str(val) -> str:
    if val is None:
        return ""
    if isinstance(val, datetime):
        return val.strftime("%d/%m/%Y")
    if isinstance(val, date):
        return val.strftime("%d/%m/%Y")
    if isinstance(val, float) and val == int(val):
        return f"{int(val)}"
    return str(val).strip()


def _celda_rango(val) -> str:
    """Texto de rango de fechas; convierte fechas Excel nativas."""
    if val is None:
        return ""
    if isinstance(val, (datetime, date)):
        return val.strftime("%d/%m/%Y")
    return _celda_str(val)


def _parsear_fila(
    fila_num: int,
    row_vals: tuple,
    cols: dict[str, int | None],
) -> FilaPlanillaArca | None:
    def get(key: str) -> str:
        i = cols[key]
        if i is None or i >= len(row_vals):
            return ""
        return _celda_str(row_vals[i])

    cuit_log = get("login")
    clave = get("clave")
    cuit_repr = get("repr")
    i_rango = cols["rango"]
    raw_rango = (
        _celda_rango(row_vals[i_rango])
        if i_rango is not None and i_rango < len(row_vals)
        else ""
    )

    def get_fecha(key: str) -> str:
        i = cols.get(key)
        if i is None or i >= len(row_vals):
            return ""
        return _celda_rango(row_vals[i])

    if not cuit_log and not clave and not cuit_repr and not raw_rango:
        return None
    if not cuit_log or not clave:
        raise CredencialesArchivoError(
            f"Fila {fila_num}: faltan CUIT de ingreso o clave fiscal."
        )
    par: tuple[str, str] | None = None
    if not raw_rango:
        fd_alt = get_fecha("desde")
        fh_alt = get_fecha("hasta")
        if fd_alt and fh_alt:
            par = (fd_alt, fh_alt)
        else:
            raise CredencialesArchivoError(
                f"Fila {fila_num}: falta el rango de fechas (columna D o columnas desde/hasta)."
            )
    else:
        par = parsear_rango_fechas_texto(raw_rango)
        if not par:
            fd_alt = get_fecha("desde")
            fh_alt = get_fecha("hasta")
            if fd_alt and fh_alt:
                par = (fd_alt, fh_alt)
    if not par:
        raise CredencialesArchivoError(
            f"Fila {fila_num}: rango de fechas inválido "
            f"(use dd/mm/yyyy - dd/mm/yyyy)."
        )
    fd, fh = par
    desde = parsear_fecha_argentina(fd)
    hasta = parsear_fecha_argentina(fh)
    validar_rango_max_un_anio(desde, hasta)

    cuit_login = _solo_digitos(cuit_log, f"CUIT ingreso fila {fila_num}")
    if cuit_repr:
        cuit_representado = _solo_digitos(cuit_repr, f"CUIT representado fila {fila_num}")
    else:
        cuit_representado = cuit_login

    return FilaPlanillaArca(
        fila_excel=fila_num,
        cuit_login=cuit_login,
        clave_fiscal=clave,
        cuit_representado=cuit_representado,
        fecha_desde=fd,
        fecha_hasta=fh,
    )


def parsear_entrada_manual(
    *,
    cuit_login: str,
    clave_fiscal: str,
    cuit_representado: str = "",
    rango_fechas: str = "",
) -> tuple[list[FilaPlanillaArca], list[str]]:
    """Convierte campos tipeados a mano en una fila de lote (sin planilla Excel)."""
    cols = {
        "login": 0,
        "clave": 1,
        "repr": 2,
        "rango": 3,
        "desde": None,
        "hasta": None,
    }
    row = (cuit_login, clave_fiscal, cuit_representado, rango_fechas)
    try:
        parsed = _parsear_fila(1, row, cols)
    except CredencialesArchivoError as exc:
        msg = str(exc).replace("Fila 1:", "Entrada manual:")
        return [], [msg]
    except Exception as exc:
        return [], [f"Entrada manual: {exc}"]
    if parsed is None:
        return [], [
            "Entrada manual: faltan CUIT ingreso, clave fiscal o período."
        ]
    return [parsed], []


def parsear_entradas_manuales(
    cuits_login: list[str],
    claves: list[str],
    cuits_repr: list[str],
    rangos: list[str],
) -> tuple[list[FilaPlanillaArca], list[str]]:
    """Convierte varias filas tipeadas a mano en filas de lote (sin Excel).

    Tolerante: una fila inválida se reporta en ``errores`` y no frena el resto.
    Las filas totalmente vacías se ignoran.
    """
    cols = {"login": 0, "clave": 1, "repr": 2, "rango": 3, "desde": None, "hasta": None}
    n = max(len(cuits_login), len(claves), len(cuits_repr), len(rangos), 0)

    def _at(lst: list[str], i: int) -> str:
        return (lst[i] if i < len(lst) else "").strip()

    filas: list[FilaPlanillaArca] = []
    errores: list[str] = []
    fila_visible = 0
    for i in range(n):
        login = _at(cuits_login, i)
        clave = _at(claves, i)
        repr_ = _at(cuits_repr, i)
        rango = _at(rangos, i)
        if not login and not clave and not repr_ and not rango:
            continue
        fila_visible += 1
        row = (login, clave, repr_, rango)
        try:
            parsed = _parsear_fila(fila_visible, row, cols)
        except CredencialesArchivoError as exc:
            errores.append(str(exc).replace(f"Fila {fila_visible}:", f"Carga manual {fila_visible}:"))
            continue
        except Exception as exc:
            errores.append(f"Carga manual {fila_visible}: {exc}")
            continue
        if parsed is not None:
            filas.append(parsed)
    return filas, errores


def leer_planilla_lote_con_errores(
    buf: io.BytesIO,
) -> tuple[list[FilaPlanillaArca], list[str]]:
    """Lee la planilla en forma tolerante.

    Devuelve ``(filas_validas, errores_por_fila)``. Una fila con CUIT o clave
    inválidos NO aborta la lectura: se reporta en ``errores_por_fila`` y se
    continúa con las siguientes. Solo se eleva excepción ante problemas del
    archivo completo (no se puede abrir o no hay ninguna fila utilizable).
    """
    try:
        buf.seek(0)
        wb = load_workbook(buf, read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
    except Exception as exc:
        raise CredencialesArchivoError("No se pudo leer la planilla Excel.") from exc

    if not rows:
        raise CredencialesArchivoError("La planilla está vacía.")

    headers = [_norm_header(c) for c in rows[0]]
    cols = _detectar_columnas(headers)
    titulos = any(
        x in h
        for h in headers
        for x in ("cuit", "clave", "rango", "fecha")
    )
    start = 2 if titulos else 1

    filas: list[FilaPlanillaArca] = []
    errores: list[str] = []
    for n, row in enumerate(rows[start - 1 :], start=start):
        try:
            parsed = _parsear_fila(n, row, cols)
        except CredencialesArchivoError as exc:
            errores.append(str(exc))
            continue
        except Exception as exc:  # fila corrupta: no frenar el resto
            errores.append(f"Fila {n}: {exc}")
            continue
        if parsed is not None:
            filas.append(parsed)

    if not filas and not errores:
        raise CredencialesArchivoError(
            "No hay filas de datos en la planilla (revisá fila 2 en adelante)."
        )
    return filas, errores


def leer_planilla_lote(buf: io.BytesIO) -> list[FilaPlanillaArca]:
    """Compatibilidad: devuelve solo las filas válidas (tolerante a filas malas)."""
    filas, _errores = leer_planilla_lote_con_errores(buf)
    if not filas:
        raise CredencialesArchivoError(
            "No hay filas válidas en la planilla (revisá CUIT, clave y fechas)."
        )
    return filas
