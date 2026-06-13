"""Lectura de entradas (Excel o manual) para «Nuestra Parte».

Columnas: CUIT ingreso, clave fiscal, CUIT representado, ejercicio.
El «ejercicio» es texto libre (p. ej. «2025» o «31/12/2025»); se usa para
elegir la opción correspondiente en el desplegable del servicio.
"""

from __future__ import annotations

import io
import re
from dataclasses import dataclass
from datetime import date, datetime

from openpyxl import load_workbook

from cuit_en_arca.errores import CredencialesArchivoError


@dataclass(frozen=True)
class FilaNuestraParte:
    fila_excel: int
    cuit_login: str
    clave_fiscal: str
    cuit_representado: str
    ejercicio: str


def _solo_digitos(s: str, etiqueta: str) -> str:
    d = re.sub(r"\D", "", str(s))
    if len(d) != 11:
        raise CredencialesArchivoError(f"{etiqueta}: se esperaban 11 dígitos.")
    return d


def _norm_header(val) -> str:
    if val is None:
        return ""
    return re.sub(r"\s+", " ", str(val).strip().lower())


def _celda_str(val) -> str:
    if val is None:
        return ""
    if isinstance(val, datetime):
        return val.strftime("%d/%m/%Y")
    if isinstance(val, date):
        return val.strftime("%d/%m/%Y")
    if isinstance(val, float) and val == int(val):
        return f"{int(val)}"
    return str(val).strip()


def _detectar_columnas(headers: list[str]) -> dict[str, int | None]:
    idx_login = idx_clave = idx_repr = idx_ej = None
    for i, h in enumerate(headers):
        if not h:
            continue
        if idx_login is None and any(
            x in h for x in ("cuit ingreso", "cuit login", "cuit representante", "cuit ingres")
        ):
            idx_login = i
        elif idx_clave is None and "clave" in h:
            idx_clave = i
        elif idx_repr is None and "representado" in h:
            idx_repr = i
        elif idx_ej is None and ("ejercicio" in h or "periodo" in h or "período" in h or "año" in h or "ano" in h):
            idx_ej = i
    if idx_login is None and idx_clave is None and idx_repr is None and idx_ej is None:
        return {"login": 0, "clave": 1, "repr": 2, "ejercicio": 3 if len(headers) > 3 else None}
    return {
        "login": idx_login if idx_login is not None else 0,
        "clave": idx_clave if idx_clave is not None else 1,
        "repr": idx_repr if idx_repr is not None else 2,
        "ejercicio": idx_ej,
    }


def _parsear_fila(fila_num: int, row_vals: tuple, cols: dict[str, int | None]) -> FilaNuestraParte | None:
    def get(key: str) -> str:
        i = cols.get(key)
        if i is None or i >= len(row_vals):
            return ""
        return _celda_str(row_vals[i])

    cuit_log = get("login")
    clave = get("clave")
    cuit_repr = get("repr")
    ejercicio = get("ejercicio")

    if not cuit_log and not clave and not cuit_repr and not ejercicio:
        return None
    if not cuit_log or not clave:
        raise CredencialesArchivoError(
            f"Fila {fila_num}: faltan CUIT de ingreso o clave fiscal."
        )

    cuit_login = _solo_digitos(cuit_log, f"CUIT ingreso fila {fila_num}")
    cuit_representado = (
        _solo_digitos(cuit_repr, f"CUIT representado fila {fila_num}")
        if cuit_repr
        else cuit_login
    )
    return FilaNuestraParte(
        fila_excel=fila_num,
        cuit_login=cuit_login,
        clave_fiscal=clave,
        cuit_representado=cuit_representado,
        ejercicio=ejercicio.strip(),
    )


def parsear_entradas_manuales_np(
    cuits_login: list[str],
    claves: list[str],
    cuits_repr: list[str],
    ejercicios: list[str],
) -> tuple[list[FilaNuestraParte], list[str]]:
    cols = {"login": 0, "clave": 1, "repr": 2, "ejercicio": 3}
    n = max(len(cuits_login), len(claves), len(cuits_repr), len(ejercicios), 0)

    def _at(lst: list[str], i: int) -> str:
        return (lst[i] if i < len(lst) else "").strip()

    filas: list[FilaNuestraParte] = []
    errores: list[str] = []
    visible = 0
    for i in range(n):
        login = _at(cuits_login, i)
        clave = _at(claves, i)
        repr_ = _at(cuits_repr, i)
        ej = _at(ejercicios, i)
        if not login and not clave and not repr_ and not ej:
            continue
        visible += 1
        try:
            parsed = _parsear_fila(visible, (login, clave, repr_, ej), cols)
        except CredencialesArchivoError as exc:
            errores.append(str(exc).replace(f"Fila {visible}:", f"Carga manual {visible}:"))
            continue
        except Exception as exc:
            errores.append(f"Carga manual {visible}: {exc}")
            continue
        if parsed is not None:
            filas.append(parsed)
    return filas, errores


def leer_planilla_np_con_errores(buf: io.BytesIO) -> tuple[list[FilaNuestraParte], list[str]]:
    try:
        buf.seek(0)
        wb = load_workbook(buf, read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
    except Exception as exc:
        raise CredencialesArchivoError("No se pudo leer la planilla Excel.") from exc

    if not rows:
        raise CredencialesArchivoError("La planilla está vacía.")

    headers = [_norm_header(c) for c in rows[0]]
    cols = _detectar_columnas(headers)
    titulos = any(x in h for h in headers for x in ("cuit", "clave", "ejercicio", "representado"))
    start = 2 if titulos else 1

    filas: list[FilaNuestraParte] = []
    errores: list[str] = []
    for n, row in enumerate(rows[start - 1 :], start=start):
        try:
            parsed = _parsear_fila(n, row, cols)
        except CredencialesArchivoError as exc:
            errores.append(str(exc))
            continue
        except Exception as exc:
            errores.append(f"Fila {n}: {exc}")
            continue
        if parsed is not None:
            filas.append(parsed)

    if not filas and not errores:
        raise CredencialesArchivoError(
            "No hay filas de datos en la planilla (revisá fila 2 en adelante)."
        )
    return filas, errores
