"""Lectura de la planilla unificada «Análisis Programado» (columnas A–G).

Solo se leen las columnas A–G; el resto de la hoja puede contener anotaciones.
Encabezados esperados (fila 1 del modelo):
  A CUIT | B Clave Fiscal | C CUIT Representado | D Fechas Mis Comprobantes
  E Fecha DFE Desde | F Fecha DFE Hasta | G Ejercicio Nuestra Parte
"""

from __future__ import annotations

import io
from dataclasses import dataclass

from openpyxl import load_workbook

from cuit_en_arca.errores import CredencialesArchivoError
from cuit_en_arca.planilla_lote import FilaPlanillaArca, _celda_rango, _celda_str, _solo_digitos
from cuit_en_arca.planilla_nuestra_parte import FilaNuestraParte
from cuit_en_arca.validacion import parsear_fecha_argentina, parsear_rango_fechas_texto, validar_rango_max_un_anio

COLS_DATOS = 7  # A–G


@dataclass(frozen=True)
class FilaAnalisisProgramado:
    fila_excel: int
    cuit_login: str
    clave_fiscal: str
    cuit_representado: str
    fechas_mis_comprobantes: str
    fecha_dfe_desde: str
    fecha_dfe_hasta: str
    ejercicio_nuestra_parte: str


def _fila_vacia(vals: tuple) -> bool:
    return not any(str(v or "").strip() for v in vals[:COLS_DATOS])


def _parsear_fila_ap(fila_num: int, row_vals: tuple) -> FilaAnalisisProgramado | None:
    vals = tuple(row_vals[:COLS_DATOS]) + ("",) * max(0, COLS_DATOS - len(row_vals))
    if _fila_vacia(vals):
        return None

    cuit_log = _celda_str(vals[0])
    clave = _celda_str(vals[1])
    cuit_repr = _celda_str(vals[2])
    fechas_mc = _celda_rango(vals[3])
    dfe_desde = _celda_rango(vals[4])
    dfe_hasta = _celda_rango(vals[5])
    ejercicio = _celda_str(vals[6])

    if not cuit_log or not clave:
        raise CredencialesArchivoError(
            f"Fila {fila_num}: faltan CUIT (col. A) o Clave Fiscal (col. B)."
        )

    cuit_login = _solo_digitos(cuit_log, f"CUIT fila {fila_num}")
    cuit_representado = (
        _solo_digitos(cuit_repr, f"CUIT representado fila {fila_num}")
        if cuit_repr
        else cuit_login
    )

    return FilaAnalisisProgramado(
        fila_excel=fila_num,
        cuit_login=cuit_login,
        clave_fiscal=clave,
        cuit_representado=cuit_representado,
        fechas_mis_comprobantes=fechas_mc.strip(),
        fecha_dfe_desde=dfe_desde.strip(),
        fecha_dfe_hasta=dfe_hasta.strip(),
        ejercicio_nuestra_parte=ejercicio.strip(),
    )


def leer_planilla_analisis_programado(
    buf: io.BytesIO,
) -> tuple[list[FilaAnalisisProgramado], list[str]]:
    try:
        buf.seek(0)
        wb = load_workbook(buf, read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
    except Exception as exc:
        raise CredencialesArchivoError("No se pudo leer la planilla Excel.") from exc

    if not rows:
        raise CredencialesArchivoError("La planilla está vacía.")

    # Fila 1 = encabezados del modelo; datos desde fila 2.
    filas: list[FilaAnalisisProgramado] = []
    errores: list[str] = []
    for n, row in enumerate(rows[1:], start=2):
        try:
            parsed = _parsear_fila_ap(n, row)
        except CredencialesArchivoError as exc:
            errores.append(str(exc))
            continue
        except Exception as exc:
            errores.append(f"Fila {n}: {exc}")
            continue
        if parsed is not None:
            filas.append(parsed)

    if not filas and not errores:
        raise CredencialesArchivoError(
            "No hay filas de datos (completá desde la fila 2, columnas A–G)."
        )
    return filas, errores


def parsear_entradas_manuales_ap(
    cuits: list[str],
    claves: list[str],
    reprs: list[str],
    fechas_mc: list[str],
    dfe_desde: list[str],
    dfe_hasta: list[str],
    ejercicios: list[str],
) -> tuple[list[FilaAnalisisProgramado], list[str]]:
    n = max(
        len(cuits),
        len(claves),
        len(reprs),
        len(fechas_mc),
        len(dfe_desde),
        len(dfe_hasta),
        len(ejercicios),
        0,
    )

    def _at(lst: list[str], i: int) -> str:
        return (lst[i] if i < len(lst) else "").strip()

    filas: list[FilaAnalisisProgramado] = []
    errores: list[str] = []
    visible = 0
    for i in range(n):
        row = (
            _at(cuits, i),
            _at(claves, i),
            _at(reprs, i),
            _at(fechas_mc, i),
            _at(dfe_desde, i),
            _at(dfe_hasta, i),
            _at(ejercicios, i),
        )
        if _fila_vacia(row):
            continue
        visible += 1
        try:
            parsed = _parsear_fila_ap(visible, row)
        except CredencialesArchivoError as exc:
            errores.append(str(exc).replace(f"Fila {visible}:", f"Carga manual {visible}:"))
            continue
        except Exception as exc:
            errores.append(f"Carga manual {visible}: {exc}")
            continue
        if parsed is not None:
            filas.append(parsed)
    return filas, errores


def _par_fechas_mc(texto: str, fila: int) -> tuple[str, str] | None:
    if not (texto or "").strip():
        return None
    par = parsear_rango_fechas_texto(texto.strip())
    if not par:
        raise CredencialesArchivoError(
            f"Fila {fila}: «Fechas Mis Comprobantes» inválido (use dd/mm/aaaa - dd/mm/aaaa)."
        )
    fd, fh = par
    desde = parsear_fecha_argentina(fd)
    hasta = parsear_fecha_argentina(fh)
    validar_rango_max_un_anio(desde, hasta)
    return fd, fh


def filas_mis_comprobantes(
    filas: list[FilaAnalisisProgramado],
) -> tuple[list[FilaPlanillaArca], list[str]]:
    out: list[FilaPlanillaArca] = []
    errores: list[str] = []
    for f in filas:
        if not f.fechas_mis_comprobantes:
            continue
        try:
            par = _par_fechas_mc(f.fechas_mis_comprobantes, f.fila_excel)
        except CredencialesArchivoError as exc:
            errores.append(str(exc))
            continue
        if not par:
            continue
        fd, fh = par
        out.append(
            FilaPlanillaArca(
                fila_excel=f.fila_excel,
                cuit_login=f.cuit_login,
                clave_fiscal=f.clave_fiscal,
                cuit_representado=f.cuit_representado,
                fecha_desde=fd,
                fecha_hasta=fh,
            )
        )
    return out, errores


@dataclass(frozen=True)
class FilaPlanillaDfeAp:
    fila_excel: int
    cuit_login: str
    clave_fiscal: str
    cuit_representado: str
    fecha_desde: str
    fecha_hasta: str


def filas_dfe(
    filas: list[FilaAnalisisProgramado],
) -> tuple[list[FilaPlanillaDfeAp], list[str]]:
    out: list[FilaPlanillaDfeAp] = []
    errores: list[str] = []
    for f in filas:
        if not f.fecha_dfe_desde and not f.fecha_dfe_hasta:
            continue
        if not f.fecha_dfe_desde or not f.fecha_dfe_hasta:
            errores.append(
                f"Fila {f.fila_excel}: DFE requiere fecha desde (col. E) y hasta (col. F)."
            )
            continue
        try:
            parsear_fecha_argentina(f.fecha_dfe_desde)
            parsear_fecha_argentina(f.fecha_dfe_hasta)
        except Exception as exc:
            errores.append(f"Fila {f.fila_excel}: fechas DFE inválidas ({exc}).")
            continue
        out.append(
            FilaPlanillaDfeAp(
                fila_excel=f.fila_excel,
                cuit_login=f.cuit_login,
                clave_fiscal=f.clave_fiscal,
                cuit_representado=f.cuit_representado,
                fecha_desde=f.fecha_dfe_desde,
                fecha_hasta=f.fecha_dfe_hasta,
            )
        )
    return out, errores


def filas_nuestra_parte(
    filas: list[FilaAnalisisProgramado],
) -> tuple[list[FilaNuestraParte], list[str]]:
    out: list[FilaNuestraParte] = []
    for f in filas:
        if not f.ejercicio_nuestra_parte:
            continue
        out.append(
            FilaNuestraParte(
                fila_excel=f.fila_excel,
                cuit_login=f.cuit_login,
                clave_fiscal=f.clave_fiscal,
                cuit_representado=f.cuit_representado,
                ejercicio=f.ejercicio_nuestra_parte,
            )
        )
    return out, []
