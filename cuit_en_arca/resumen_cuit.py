"""Genera el Excel resumen por CUIT con tabla plana e hipervínculos al detalle mensual.

Hoja «Resumen»: una fila por CUIT con columnas MCE y MCR en secuencia.
Cada importe enlaza a una hoja de detalle con la composición mensual del valor.
"""

from __future__ import annotations

import io
import re
from dataclasses import dataclass, field

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

RESUMEN_HEADERS = [
    "CUIT",
    "Razón Social",
    "Neto gravado total (MCE)",
    "No gravado + exento (MCE)",
    "Total IVA (MCE)",
    "Total (MCE)",
    "Neto gravado total (MCR)",
    "Total IVA (MCR)",
    "Total (MCR)",
]

NOMBRE_SALIDA = "Resumen por CUIT.xlsx"

_FORMATO_CONTABILIDAD = (
    r'_ * #,##0.00_ ;_ * (#,##0.00)_ ;_ * "-"??_ ;_ @_ '
)
_FONT_HIPERVINCULO = Font(color="0563C1", underline="single")


@dataclass
class _AcumTipo:
    # mes "YYYY-MM" -> [neto, no_grav_exento, iva, total]
    por_mes: dict[str, list[float]] = field(default_factory=dict)

    def sumar(
        self,
        mes: str,
        neto: float,
        no_grav_exento: float,
        iva: float,
        total: float,
    ) -> None:
        acc = self.por_mes.setdefault(mes, [0.0, 0.0, 0.0, 0.0])
        acc[0] += neto
        acc[1] += no_grav_exento
        acc[2] += iva
        acc[3] += total

    def totales(self) -> tuple[float, float, float, float]:
        neto = no_grav = iva = total = 0.0
        for vals in self.por_mes.values():
            neto += vals[0]
            no_grav += vals[1]
            iva += vals[2]
            total += vals[3]
        return neto, no_grav, iva, total


@dataclass
class _AcumCuit:
    cuit: str
    razon_social: str = ""
    emitidos: _AcumTipo = field(default_factory=_AcumTipo)
    recibidos: _AcumTipo = field(default_factory=_AcumTipo)


class ResumenCuitAcumulador:
    """Acumula totales por CUIT/tipo/mes para el Excel resumen."""

    def __init__(self) -> None:
        self._cuits: dict[str, _AcumCuit] = {}

    def agregar(
        self,
        cuit: str,
        *,
        emitidos: bool,
        razon_social: str,
        por_mes: dict[str, dict[str, float]],
    ) -> None:
        ac = self._cuits.setdefault(cuit, _AcumCuit(cuit=cuit))
        if razon_social and not ac.razon_social:
            ac.razon_social = razon_social
        destino = ac.emitidos if emitidos else ac.recibidos
        for mes, vals in por_mes.items():
            destino.sumar(
                mes,
                float(vals.get("neto", 0.0)),
                float(vals.get("no_grav_exento", 0.0)),
                float(vals.get("iva", 0.0)),
                float(vals.get("total", 0.0)),
            )

    def tiene_datos(self) -> bool:
        return any(
            ac.emitidos.por_mes or ac.recibidos.por_mes
            for ac in self._cuits.values()
        )

    def cuits_ordenados(self) -> list[_AcumCuit]:
        return [self._cuits[cuit] for cuit in sorted(self._cuits)]


def _formatear_cuit(cuit: str) -> str:
    digitos = re.sub(r"\D", "", cuit)
    if len(digitos) == 11:
        return f"{digitos[:2]}-{digitos[2:10]}-{digitos[10]}"
    return str(cuit)


def _nombre_hoja_detalle(cuit: str) -> str:
    digitos = re.sub(r"\D", "", cuit) or cuit
    return f"Det {digitos}"[:31]


def _aplicar_contabilidad(celda) -> None:
    celda.number_format = _FORMATO_CONTABILIDAD


def _escribir_fila_mensual(
    ws,
    row: int,
    mes: str,
    valores: list[float],
    *,
    col_inicio: int = 1,
) -> None:
    ws.cell(row, col_inicio, mes)
    for offset, valor in enumerate(valores, start=1):
        celda = ws.cell(row, col_inicio + offset, round(float(valor), 2))
        _aplicar_contabilidad(celda)


def _escribir_total_seccion(
    ws,
    row: int,
    first_data_row: int,
    last_data_row: int,
    *,
    col_inicio: int,
    n_columnas: int,
) -> None:
    if last_data_row < first_data_row:
        return
    ws.cell(row, col_inicio, "Total")
    for offset in range(1, n_columnas + 1):
        col = col_inicio + offset
        col_l = get_column_letter(col)
        celda = ws.cell(
            row,
            col,
            f"=SUM({col_l}{first_data_row}:{col_l}{last_data_row})",
        )
        _aplicar_contabilidad(celda)


def _escribir_hoja_detalle(ws, ac: _AcumCuit) -> dict[str, str]:
    """Detalle mensual del CUIT. Devuelve celdas destino para hipervínculos."""
    anchors: dict[str, str] = {}
    titulo = _formatear_cuit(ac.cuit)
    if ac.razon_social:
        titulo = f"{titulo} — {ac.razon_social}"
    ws["A1"] = titulo

    row = 3
    ws.cell(row, 1, "Mis Comprobantes Emitidos")
    row += 1
    headers_emitidos = [
        "Mes",
        "Neto Gravado Total",
        "Ingresos no grav. y exentos",
        "Total IVA",
        "Imp. Total",
    ]
    for j, encab in enumerate(headers_emitidos, start=1):
        ws.cell(row, j, encab)
    header_emitidos = row
    row += 1
    first_emitidos = row
    for mes in sorted(ac.emitidos.por_mes):
        neto, no_grav, iva, total = ac.emitidos.por_mes[mes]
        _escribir_fila_mensual(ws, row, mes, [neto, no_grav, iva, total])
        row += 1
    if row > first_emitidos:
        _escribir_total_seccion(
            ws,
            row,
            first_emitidos,
            row - 1,
            col_inicio=1,
            n_columnas=4,
        )
        row += 1

    anchors["mce_neto"] = f"B{header_emitidos}"
    anchors["mce_no_grav"] = f"C{header_emitidos}"
    anchors["mce_iva"] = f"D{header_emitidos}"
    anchors["mce_total"] = f"E{header_emitidos}"

    row += 1
    ws.cell(row, 1, "Mis Comprobantes Recibidos")
    row += 1
    headers_recibidos = ["Mes", "Neto Gravado Total", "Total IVA", "Imp. Total"]
    for j, encab in enumerate(headers_recibidos, start=1):
        ws.cell(row, j, encab)
    header_recibidos = row
    row += 1
    first_recibidos = row
    for mes in sorted(ac.recibidos.por_mes):
        neto, _no_grav, iva, total = ac.recibidos.por_mes[mes]
        _escribir_fila_mensual(ws, row, mes, [neto, iva, total])
        row += 1
    if row > first_recibidos:
        _escribir_total_seccion(
            ws,
            row,
            first_recibidos,
            row - 1,
            col_inicio=1,
            n_columnas=3,
        )

    anchors["mcr_neto"] = f"B{header_recibidos}"
    anchors["mcr_iva"] = f"C{header_recibidos}"
    anchors["mcr_total"] = f"D{header_recibidos}"
    return anchors


def _celda_resumen_con_detalle(
    ws,
    row: int,
    col: int,
    valor: float,
    hoja_detalle: str,
    celda_detalle: str,
) -> None:
    celda = ws.cell(row, col, round(float(valor), 2))
    _aplicar_contabilidad(celda)
    destino = f"'{hoja_detalle}'!{celda_detalle}"
    celda.hyperlink = f"#{destino}"
    celda.font = _FONT_HIPERVINCULO


def construir_resumen_cuit_xlsx(acumulador: ResumenCuitAcumulador) -> bytes | None:
    """Devuelve el xlsx resumen (bytes) o None si no hay datos."""
    if not acumulador.tiene_datos():
        return None

    wb = Workbook()
    ws_resumen = wb.active
    ws_resumen.title = "Resumen"

    for col, encab in enumerate(RESUMEN_HEADERS, start=1):
        ws_resumen.cell(1, col, encab)

    fila = 2
    for ac in acumulador.cuits_ordenados():
        hoja_det = _nombre_hoja_detalle(ac.cuit)
        ws_det = wb.create_sheet(hoja_det)
        anchors = _escribir_hoja_detalle(ws_det, ac)

        neto_e, no_grav_e, iva_e, total_e = ac.emitidos.totales()
        neto_r, _no_grav_r, iva_r, total_r = ac.recibidos.totales()

        ws_resumen.cell(fila, 1, _formatear_cuit(ac.cuit))
        ws_resumen.cell(fila, 2, ac.razon_social or _formatear_cuit(ac.cuit))

        columnas_valores = (
            (3, neto_e, "mce_neto"),
            (4, no_grav_e, "mce_no_grav"),
            (5, iva_e, "mce_iva"),
            (6, total_e, "mce_total"),
            (7, neto_r, "mcr_neto"),
            (8, iva_r, "mcr_iva"),
            (9, total_r, "mcr_total"),
        )
        for col, valor, clave in columnas_valores:
            destino = anchors.get(clave)
            if destino:
                _celda_resumen_con_detalle(
                    ws_resumen, fila, col, valor, hoja_det, destino
                )
            else:
                celda = ws_resumen.cell(fila, col, round(float(valor), 2))
                _aplicar_contabilidad(celda)

        fila += 1

    salida = io.BytesIO()
    wb.save(salida)
    return salida.getvalue()
